"""
excel/sheet_depreciation.py
══════════════════════════════
WDV (Written Down Value) depreciation per Income Tax Act.
One block per asset class: opening WDV, annual charge, closing WDV.
Net Fixed Assets row used by Expenses and BS.
"""

from __future__ import annotations
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore, AssetCategory
from excel.sheet_costmeans import cat_summary_row
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_REF,
    FILL_TOTAL, AL_LEFT, AL_CENTER,
    FMT_LAKHS, FMT_PCT,
    font, fill
)


# Map AssetCategory → layout engine asset type key
ASSET_CAT_MAP = {
    AssetCategory.PLANT_MACHINERY: "plant_machinery",
    AssetCategory.CIVIL_WORKS:     "civil_works",
    AssetCategory.FURNITURE:       "furniture",
    AssetCategory.VEHICLE:         "vehicle",
    AssetCategory.ELECTRICAL:      "electrical",
    AssetCategory.PRE_OPERATIVE:   "other",
    AssetCategory.OTHER:           "other",
}

DEPR_KEY_MAP = {
    "plant_machinery": "depr.plant_machinery",
    "civil_works":     "depr.civil_works",
    "furniture":       "depr.furniture",
    "vehicle":         "depr.vehicle",
    "electrical":      "depr.electrical",
    "other":           "depr.other",
}


def write_depreciation_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("Depreciation")

    set_col_widths(ws, {
        "A": 4, "B": 36, "C": 14,
        **{get_column_letter(layout.DEP_COL_YEAR1 + y - 1): 13
           for y in range(1, layout.n_years + 1)}
    })

    write_title(ws, store.project_profile.company_name,
                "Depreciation Schedule (IT Act — WDV Method)")

    # Year headers
    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Asset / Particulars").font = font(bold=True)
    ws.cell(row=4, column=3, value="Rate").font = font(bold=True, size=9, color="808080")
    for y in range(1, layout.n_years + 1):
        col = layout.DEP_COL_YEAR1 + y - 1
        c = ws.cell(row=4, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    # Group assets by category
    assets_by_cat: dict[str, float] = {}
    for asset in store.capital_means.assets:
        cat_key = ASSET_CAT_MAP.get(asset.category, "other")
        assets_by_cat[cat_key] = assets_by_cat.get(cat_key, 0.0) + asset.cost_lakhs

    present_cats = [c for c in LayoutEngine.ASSET_ORDER if c in assets_by_cat]

    for cat_key in present_cats:
        cost = assets_by_cat[cat_key]
        depr_key = DEPR_KEY_MAP[cat_key]
        depr_ref = layout.asmp_ref(depr_key)

        opening_row = layout.dep_asset_row(cat_key, "opening")
        depr_row    = layout.dep_asset_row(cat_key, "depr")
        closing_row = layout.dep_asset_row(cat_key, "closing")

        label = cat_key.replace("_", " ").title()
        ws.cell(row=opening_row, column=2, value=f"{label} — Opening WDV").font = FONT_BODY
        ws.cell(row=depr_row,    column=2, value=f"  Less: Depreciation").font = FONT_BODY
        ws.cell(row=closing_row, column=2, value=f"  Closing WDV").font = font(bold=True)
        ws.cell(row=depr_row,    column=3,
                value=f"={depr_ref}").font = FONT_REF
        ws.cell(row=depr_row,    column=3).number_format = FMT_PCT

        # Year 1 opening WDV = category total from Cost & Means sheet (not hardcoded)
        cm_row = cat_summary_row(cat_key)
        cm_ref = f"'Cost & Means'!$D${cm_row}"

        for y in range(1, layout.n_years + 1):
            col   = layout.DEP_COL_YEAR1 + y - 1
            col_l = get_column_letter(col)

            if y == 1:
                # Year 1 opening = category total from Cost & Means
                ws.cell(row=opening_row, column=col,
                        value=f"={cm_ref}").font = FONT_FORMULA
            else:
                prev_col = get_column_letter(col - 1)
                ws.cell(row=opening_row, column=col,
                        value=f"={prev_col}{closing_row}").font = FONT_FORMULA

            # Depreciation = opening × rate (half-year rule: if months <=6, use rate/2)
            # impl months used: if impl_months <= 6 in year 1, apply half rate
            impl_ref = layout.asmp_ref("impl.months")
            if y == 1:
                ws.cell(row=depr_row, column=col,
                        value=(f"=IF({impl_ref}<=6,"
                               f"{col_l}{opening_row}*{depr_ref}/2,"
                               f"{col_l}{opening_row}*{depr_ref})")).font = FONT_FORMULA
            else:
                ws.cell(row=depr_row, column=col,
                        value=f"={col_l}{opening_row}*{depr_ref}").font = FONT_FORMULA

            # Closing = opening - depreciation
            ws.cell(row=closing_row, column=col,
                    value=f"={col_l}{opening_row}-{col_l}{depr_row}").font = FONT_FORMULA

            for row in [opening_row, depr_row, closing_row]:
                ws.cell(row=row, column=col).number_format = FMT_LAKHS

    # ── Total depreciation charge row ─────────────────────────────────────────
    total_depr_row = layout.dep_total_depr_row()
    ws.row_dimensions[total_depr_row].height = 17
    t = ws.cell(row=total_depr_row, column=2,
                value="TOTAL DEPRECIATION CHARGE")
    t.font = font(bold=True, color="FFFFFF")
    t.fill = fill("2E75B6")
    ws.cell(row=total_depr_row, column=3, value="₹ Lakhs").font = font(bold=True, size=9)

    for y in range(1, layout.n_years + 1):
        col   = layout.DEP_COL_YEAR1 + y - 1
        col_l = get_column_letter(col)
        depr_rows = [layout.dep_asset_row(c, "depr") for c in present_cats]
        total = "+".join(f"{col_l}{r}" for r in depr_rows)
        c = ws.cell(row=total_depr_row, column=col, value=f"={total}")
        c.font          = font(bold=True, color="FFFFFF")
        c.fill          = fill("2E75B6")
        c.number_format = FMT_LAKHS

    # ── Net Fixed Assets (WDV) row ─────────────────────────────────────────────
    nfa_row = layout.dep_net_block_row()
    ws.row_dimensions[nfa_row].height = 17
    n = ws.cell(row=nfa_row, column=2, value="NET FIXED ASSETS (WDV)")
    n.font = font(bold=True, color="1F3864")
    n.fill = fill("EBF5FB")
    ws.cell(row=nfa_row, column=3, value="₹ Lakhs").font = font(bold=True, size=9)

    for y in range(1, layout.n_years + 1):
        col   = layout.DEP_COL_YEAR1 + y - 1
        col_l = get_column_letter(col)
        closing_rows = [layout.dep_asset_row(c, "closing") for c in present_cats]
        total = "+".join(f"{col_l}{r}" for r in closing_rows)
        c = ws.cell(row=nfa_row, column=col, value=f"={total}")
        c.font          = font(bold=True, color="1F3864")
        c.fill          = fill("EBF5FB")
        c.number_format = FMT_LAKHS

    freeze_header(ws, row=5, col=4)
    return ws
