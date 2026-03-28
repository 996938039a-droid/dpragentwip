"""
excel/styles.py
════════════════
Shared styles, fonts, fills used across all sheet writers.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

FONT_NAME = "Arial"

def font(bold=False, size=10, color="000000", italic=False):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color, italic=italic)

def fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def border(sides="bottom", color="CCCCCC", style="thin"):
    s = Side(style=style, color=color)
    kwargs = {side: s for side in sides.split(",")}
    return Border(**kwargs)

def full_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

# ── Standard fills ────────────────────────────────────────────────────────────
FILL_HEADER    = fill("1F3864")   # dark navy
FILL_SUBHEADER = fill("D6DCE4")   # light grey-blue
FILL_TOTAL     = fill("E8F0FE")   # very light blue for totals
FILL_ALT       = fill("F8F9FA")   # alternating row
FILL_WHITE     = fill("FFFFFF")

# ── Standard fonts ────────────────────────────────────────────────────────────
FONT_HEADER    = font(bold=True, color="FFFFFF", size=10)
FONT_SUBHEADER = font(bold=True, color="1F3864", size=10)
FONT_TOTAL     = font(bold=True, color="1F3864", size=10)
FONT_BODY      = font(size=10)
FONT_FORMULA   = font(color="000080", size=10)   # dark blue for formula cells
FONT_REF       = font(color="006400", size=10)   # dark green for cross-sheet refs

# ── Alignments ────────────────────────────────────────────────────────────────
AL_LEFT   = Alignment(horizontal="left",   vertical="center")
AL_RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=False)
AL_CENTER = Alignment(horizontal="center", vertical="center")

# ── Number formats ────────────────────────────────────────────────────────────
FMT_LAKHS   = '#,##0.00'        # monetary in lakhs
FMT_INT     = '#,##0'           # whole numbers
FMT_PCT     = '0.00%'           # percentage
FMT_DATE    = 'DD-MMM-YYYY'     # date
FMT_ZERO    = '#,##0.00;(#,##0.00);"-"'  # zero shows as dash

def write_header_row(ws, row: int, cols: list[tuple], height: int = 20):
    """Write a navy header row. cols = [(col_idx, text), ...]"""
    ws.row_dimensions[row].height = height
    for col, text in cols:
        c = ws.cell(row=row, column=col, value=text)
        c.font      = FONT_HEADER
        c.fill      = FILL_HEADER
        c.alignment = AL_CENTER

def write_title(ws, company_name: str, sheet_title: str):
    """Write standard 2-line title block."""
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 14
    t = ws.cell(row=1, column=2, value=sheet_title.upper())
    t.font = font(bold=True, size=12, color="1F3864")
    s = ws.cell(row=2, column=2,
                value=f"{company_name}  |  (All amounts in INR Lakhs unless otherwise stated)")
    s.font = font(size=9, color="808080", italic=True)

def set_col_widths(ws, widths: dict[str, float]):
    """Set column widths. widths = {"A": 8, "B": 40, ...}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze_header(ws, row: int = 5, col: int = 3):
    """Freeze panes at row/col for easy scrolling."""
    from openpyxl.utils import get_column_letter
    ws.freeze_panes = f"{get_column_letter(col)}{row}"
