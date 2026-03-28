"""
excel/sheet_tax.py
════════════════════
Tax computation for Company / LLP / Proprietorship.
Entity type drives which tax table is used (formula-based selection).
"""

from __future__ import annotations
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore, EntityType
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_REF,
    AL_LEFT, AL_CENTER, FMT_LAKHS, FMT_PCT,
    font, fill
)


def write_tax_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("Tax")

    set_col_widths(ws, {
        "A": 4, "B": 36, "C": 14,
        **{get_column_letter(layout.year_col(y)): 13
           for y in range(1, layout.n_years + 1)}
    })

    write_title(ws, store.project_profile.company_name, "Tax Computation")

    # Year headers
    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Particulars").font = font(bold=True)
    for y in range(1, layout.n_years + 1):
        col = layout.year_col(y)
        c = ws.cell(row=4, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    entity_ref = layout.asmp_ref("tax.entity_type")
    rate_ref   = layout.asmp_ref("tax.company_rate")
    sur_ref    = layout.asmp_ref("tax.surcharge")
    hec_ref    = layout.asmp_ref("tax.hec")

    # ── Tax computation rows ───────────────────────────────────────────────────
    PBT_REF_ROW  = 6   # where we write the PBT reference
    BASIC_ROW    = 8
    SURCH_ROW    = 9
    HEC_ROW      = 10
    TOTAL_ROW    = 11  # this is what PL reads

    ws.cell(row=PBT_REF_ROW, column=2,
            value="Profit Before Tax (from P&L)").font = FONT_BODY
    ws.cell(row=7, column=2,
            value="[Applied rate depends on entity type]").font = font(size=9, italic=True, color="808080")
    ws.cell(row=BASIC_ROW, column=2,
            value="  Basic Tax").font = FONT_BODY
    ws.cell(row=SURCH_ROW, column=2,
            value="  Surcharge").font = FONT_BODY
    ws.cell(row=HEC_ROW, column=2,
            value="  Health & Education Cess").font = FONT_BODY
    ws.cell(row=TOTAL_ROW, column=2,
            value="TOTAL TAX").font = FONT_TOTAL

    for y in range(1, layout.n_years + 1):
        col   = layout.year_col(y)
        col_l = layout.year_col_letter(y)
        pl_col = layout.pl_col_letter(y)

        # PBT reference — reads from PL sheet
        ws.cell(row=PBT_REF_ROW, column=col,
                value=f"=PL!{pl_col}{layout.PL_PBT_ROW}").font = FONT_REF
        ws.cell(row=PBT_REF_ROW, column=col).number_format = FMT_LAKHS

        # Basic tax — all rates referenced from Assumption sheet, not hardcoded
        llp_ref = layout.asmp_ref("tax.llp_rate")
        ws.cell(row=BASIC_ROW, column=col,
                value=(f"=IF({entity_ref}=\"Company\","
                       f"MAX({col_l}{PBT_REF_ROW},0)*{rate_ref},"
                       f"IF({entity_ref}=\"LLP\","
                       f"MAX({col_l}{PBT_REF_ROW},0)*{llp_ref},"
                       f"MAX({col_l}{PBT_REF_ROW},0)*{llp_ref}))")).font = FONT_FORMULA
        ws.cell(row=BASIC_ROW, column=col).number_format = FMT_LAKHS

        # Surcharge (only if PBT > threshold — from Assumption, not hardcoded)
        thresh_ref = layout.asmp_ref("tax.surcharge_threshold")
        ws.cell(row=SURCH_ROW, column=col,
                value=(f"=IF({col_l}{PBT_REF_ROW}>{thresh_ref},"
                       f"{col_l}{BASIC_ROW}*{sur_ref},0)")).font = FONT_FORMULA
        ws.cell(row=SURCH_ROW, column=col).number_format = FMT_LAKHS

        # HEC = (Basic + Surcharge) × 4%
        ws.cell(row=HEC_ROW, column=col,
                value=f"=({col_l}{BASIC_ROW}+{col_l}{SURCH_ROW})*{hec_ref}").font = FONT_FORMULA
        ws.cell(row=HEC_ROW, column=col).number_format = FMT_LAKHS

        # Total = basic + surcharge + HEC, but zero if PBT <= 0
        ws.cell(row=TOTAL_ROW, column=col,
                value=(f"=IF({col_l}{PBT_REF_ROW}<=0,0,"
                       f"{col_l}{BASIC_ROW}+{col_l}{SURCH_ROW}+{col_l}{HEC_ROW})")).font = FONT_TOTAL
        ws.cell(row=TOTAL_ROW, column=col).number_format = FMT_LAKHS
        ws.cell(row=TOTAL_ROW, column=col).fill = fill("EBF5FB")

    return ws
