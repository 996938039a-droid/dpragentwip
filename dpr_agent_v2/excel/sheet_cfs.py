"""excel/sheet_cfs.py — Cash Flow Statement with OD, Drawings, Investments, Non-Op Income."""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FMT_LAKHS,
    font, fill
)

def write_cfs_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("CFS")
    set_col_widths(ws, {"A": 4, "B": 42,
                        **{get_column_letter(layout.cfs_col(y)): 13 for y in range(1, layout.n_years+1)}})
    write_title(ws, store.project_profile.company_name, "Projected Cash Flow Statement  (Indirect Method)")

    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Particulars").font = font(bold=True)
    for y in range(1, layout.n_years+1):
        col = layout.cfs_col(y)
        c = ws.cell(row=4, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")

    def lbl(row, text, bold=False):
        ws.cell(row=row, column=2, value=text).font = font(bold=bold)

    def sec(row, text):
        c = ws.cell(row=row, column=2, value=text)
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("2E75B6")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+layout.n_years+2)

    def fr(row, fn, bold=False, bg=None):
        for y in range(1, layout.n_years+1):
            col = layout.cfs_col(y); col_l = get_column_letter(col)
            c = ws.cell(row=row, column=col, value=fn(y, col_l))
            c.font = font(bold=bold); c.number_format = FMT_LAKHS
            if bg: c.fill = fill(bg)

    pl = layout.pl_col_letter
    wc = lambda y: get_column_letter(layout.wcap_col(y))
    pwc = lambda y: get_column_letter(layout.wcap_col(y-1)) if y > 1 else None
    EXP = LayoutEngine

    # ── A. Operating ──────────────────────────────────────────────────────────
    sec(layout.CFS_PBT_ROW-1, "A.  CASH FROM OPERATING ACTIVITIES")
    lbl(layout.CFS_PBT_ROW,     "  Profit Before Tax")
    lbl(layout.CFS_DEPR_ROW,    "  Add: Depreciation")
    lbl(layout.CFS_INTEREST_ROW,"  Add: Finance Costs (Interest)")
    fr(layout.CFS_PBT_ROW,      lambda y,c: f"=PL!{pl(y)}{layout.PL_PBT_ROW}")
    fr(layout.CFS_DEPR_ROW,     lambda y,c: f"=PL!{pl(y)}{layout.PL_DEPR_ROW}")
    fr(layout.CFS_INTEREST_ROW, lambda y,c: f"=PL!{pl(y)}{layout.PL_TOTAL_FINANCE_ROW}")

    # WC changes (rows 9, 10, 11)
    lbl(layout.CFS_WC_CHANGES_ROW,   "    (Increase)/Decrease in Debtors")
    lbl(layout.CFS_WC_CHANGES_ROW+1, "    (Increase)/Decrease in Stock")
    lbl(layout.CFS_WC_CHANGES_ROW+2, "    Increase/(Decrease) in Trade Creditors")
    for y in range(1, layout.n_years+1):
        col = layout.cfs_col(y); col_l = get_column_letter(col)
        wcol = wc(y); prev = pwc(y)
        # Debtors
        d_f = f"=-'W Cap'!{wcol}{layout.WCAP_DEBTORS}" if y==1 else f"='W Cap'!{prev}{layout.WCAP_DEBTORS}-'W Cap'!{wcol}{layout.WCAP_DEBTORS}"
        ws.cell(row=layout.CFS_WC_CHANGES_ROW, column=col, value=d_f).font = FONT_FORMULA
        ws.cell(row=layout.CFS_WC_CHANGES_ROW, column=col).number_format = FMT_LAKHS
        # Stock
        s_f = f"=-'W Cap'!{wcol}{layout.WCAP_STOCK_RM}" if y==1 else f"='W Cap'!{prev}{layout.WCAP_STOCK_RM}-'W Cap'!{wcol}{layout.WCAP_STOCK_RM}"
        ws.cell(row=layout.CFS_WC_CHANGES_ROW+1, column=col, value=s_f).font = FONT_FORMULA
        ws.cell(row=layout.CFS_WC_CHANGES_ROW+1, column=col).number_format = FMT_LAKHS
        # Creditors
        cr_f = f"='W Cap'!{wcol}{layout.WCAP_TOTAL_CL}" if y==1 else f"='W Cap'!{wcol}{layout.WCAP_TOTAL_CL}-'W Cap'!{prev}{layout.WCAP_TOTAL_CL}"
        ws.cell(row=layout.CFS_WC_CHANGES_ROW+2, column=col, value=cr_f).font = FONT_FORMULA
        ws.cell(row=layout.CFS_WC_CHANGES_ROW+2, column=col).number_format = FMT_LAKHS

    lbl(layout.CFS_TAX_ROW,      "    Less: Taxes Paid")
    fr(layout.CFS_TAX_ROW,       lambda y,c: f"=-PL!{pl(y)}{layout.PL_TAX_ROW}")

    # Drawings (new row 13)
    drawings_ref = layout.asmp_ref("opex.drawings_base")
    drawings_esc = layout.asmp_ref("opex.drawings_escalation")
    lbl(layout.CFS_DRAWINGS_ROW, "    Less: Drawings / Proprietor Withdrawals")
    fr(layout.CFS_DRAWINGS_ROW,
       lambda y,c: (f"=-{drawings_ref}*(1+{drawings_esc})^{y-1}" if y > 1 else f"=-{drawings_ref}"))

    lbl(layout.CFS_NET_OPERATING, "NET CASH FROM OPERATIONS", bold=True)
    fr(layout.CFS_NET_OPERATING,
       lambda y,c: f"=SUM({c}{layout.CFS_PBT_ROW}:{c}{layout.CFS_DRAWINGS_ROW})",
       bold=True, bg="EBF5FB")

    # ── B. Investing ──────────────────────────────────────────────────────────
    sec(layout.CFS_CAPEX_ROW-1, "B.  CASH FROM INVESTING ACTIVITIES")

    from excel.sheet_costmeans import ASSET_DATA_START
    n_assets = len(store.capital_means.assets)
    cm_total_row = ASSET_DATA_START + n_assets + 1

    # Capex
    lbl(layout.CFS_CAPEX_ROW, "  Capital Expenditure (Fixed Assets)")
    fr(layout.CFS_CAPEX_ROW, lambda y,c: (f"=-'Cost & Means'!$D${cm_total_row}" if y==1 else "=0"))

    # Investment & Deposits
    fd_ref = layout.asmp_ref("wc.investment_deposits")
    lbl(layout.CFS_FD_ROW, "  Investment & Fixed Deposits")
    fr(layout.CFS_FD_ROW, lambda y,c: (f"=-{fd_ref}" if y==1 else "=0"))

    # Other Non-Current Assets
    nc_ref         = layout.asmp_ref("wc.other_non_current")
    intangible_ref = layout.asmp_ref("balance.intangible_assets")
    nc_inv_ref     = layout.asmp_ref("balance.nc_investments")
    sec_dep_ref    = layout.asmp_ref("balance.security_deposits")
    lbl(layout.CFS_OTHER_NC_ROW, "  Other Non-Current Assets  (Intangibles + NC Investments + Security Deps)")
    fr(layout.CFS_OTHER_NC_ROW,
       lambda y,c: (f"=-({nc_ref}+{intangible_ref}+{nc_inv_ref}+{sec_dep_ref})" if y==1 else "=0"))

    # Non-Operating Income
    noi_ref = layout.asmp_ref("wc.non_operating_income")
    lbl(layout.CFS_NON_OP_ROW, "  Non-Operating Income")
    fr(layout.CFS_NON_OP_ROW, lambda y,c: f"={noi_ref}")

    lbl(layout.CFS_NET_INVESTING, "NET CASH FROM INVESTING", bold=True)
    fr(layout.CFS_NET_INVESTING,
       lambda y,c: f"=SUM({c}{layout.CFS_CAPEX_ROW}:{c}{layout.CFS_NON_OP_ROW})",
       bold=True, bg="EBF5FB")

    # ── C. Financing ──────────────────────────────────────────────────────────
    sec(layout.CFS_TL_ROW-1, "C.  CASH FROM FINANCING ACTIVITIES")

    # Equity row in Cost & Means
    fin_data_start = ASSET_DATA_START + n_assets + 1 + 3 + 1
    equity_row = None
    for i, src in enumerate(store.capital_means.finance_sources):
        if src.is_equity:
            equity_row = fin_data_start + i; break

    lbl(layout.CFS_TL_ROW,     "  Promoter Equity Brought In")
    lbl(layout.CFS_TL_ROW+1,   "  Term Loan Drawdown / (Repayment)")
    lbl(layout.CFS_OD_ROW,     "  OD / Working Capital Loan")
    lbl(layout.CFS_INTEREST_PAID, "  Finance Costs Paid")

    fr(layout.CFS_TL_ROW,
       lambda y,c: (f"='Cost & Means'!$D${equity_row}" if (y==1 and equity_row) else "=0"))

    tl_amt_ref = layout.asmp_ref("finance.tl_amount")
    fr(layout.CFS_TL_ROW+1,
       lambda y,c: (f"={tl_amt_ref}-'Term Loan'!{get_column_letter(EXP.TL_ANNUAL_COL_Y1+y-1)}{EXP.TL_ANNUAL_PRINCIPAL_ROW}"
                    if y==1
                    else f"=-'Term Loan'!{get_column_letter(EXP.TL_ANNUAL_COL_Y1+y-1)}{EXP.TL_ANNUAL_PRINCIPAL_ROW}"))

    # OD drawdown: Year 1 = full OD limit, subsequent years = change in OD balance
    od_ref = layout.asmp_ref("wc.wc_loan_amount")
    for y in range(1, layout.n_years+1):
        col = layout.cfs_col(y); col_l = get_column_letter(col)
        wcol = wc(y); prev = pwc(y)
        if y == 1:
            # Year 1: draw down the OD limit
            od_f = f"={od_ref}"
        else:
            # Subsequent years: OD is fixed (sanctioned limit), change = 0
            od_f = "=0"
        ws.cell(row=layout.CFS_OD_ROW, column=col, value=od_f).font = FONT_FORMULA
        ws.cell(row=layout.CFS_OD_ROW, column=col).number_format = FMT_LAKHS

    fr(layout.CFS_INTEREST_PAID, lambda y,c: f"=-PL!{pl(y)}{layout.PL_TOTAL_FINANCE_ROW}")

    lbl(layout.CFS_NET_FINANCING, "NET CASH FROM FINANCING", bold=True)
    fr(layout.CFS_NET_FINANCING,
       lambda y,c: (f"={c}{layout.CFS_TL_ROW}+{c}{layout.CFS_TL_ROW+1}"
                    f"+{c}{layout.CFS_OD_ROW}+{c}{layout.CFS_INTEREST_PAID}"),
       bold=True, bg="EBF5FB")

    # ── Net change & closing ─────────────────────────────────────────────────
    lbl(layout.CFS_NET_CHANGE,   "NET CHANGE IN CASH", bold=True)
    lbl(layout.CFS_OPENING_CASH, "  Opening Cash Balance")
    lbl(layout.CFS_CLOSING_CASH, "CLOSING CASH BALANCE", bold=True)

    fr(layout.CFS_NET_CHANGE,
       lambda y,c: f"={c}{layout.CFS_NET_OPERATING}+{c}{layout.CFS_NET_INVESTING}+{c}{layout.CFS_NET_FINANCING}",
       bold=True)

    for y in range(1, layout.n_years+1):
        col = layout.cfs_col(y); col_l = get_column_letter(col)
        if y == 1:
            ws.cell(row=layout.CFS_OPENING_CASH, column=col, value=0)
        else:
            prev_l = get_column_letter(layout.cfs_col(y-1))
            ws.cell(row=layout.CFS_OPENING_CASH, column=col, value=f"={prev_l}{layout.CFS_CLOSING_CASH}").font = FONT_FORMULA
        ws.cell(row=layout.CFS_OPENING_CASH, column=col).number_format = FMT_LAKHS
        ws.cell(row=layout.CFS_CLOSING_CASH, column=col, value=f"={col_l}{layout.CFS_OPENING_CASH}+{col_l}{layout.CFS_NET_CHANGE}").font = FONT_FORMULA
        ws.cell(row=layout.CFS_CLOSING_CASH, column=col).number_format = FMT_LAKHS
        ws.cell(row=layout.CFS_CLOSING_CASH, column=col).fill = fill("EBF5FB")

    freeze_header(ws, row=5, col=layout.CFS_COL_YEAR1)
    return ws
