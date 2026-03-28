"""excel/sheet_bs.py — Balance Sheet: all line items from Assumption/W Cap/CFS (nothing hardcoded 0)."""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL,
    FMT_LAKHS, FMT_PCT,
    font, fill
)

def write_bs_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("BS")
    set_col_widths(ws, {"A": 4, "B": 42,
                        **{get_column_letter(layout.bs_col(y)): 13 for y in range(1, layout.n_years+1)}})

    ws.cell(row=1, column=2, value="Balance Sheet").font = font(bold=True, size=14, color="1F3864")
    ws.cell(row=2, column=2, value=store.project_profile.company_name).font = font(bold=True, size=11)
    ws.cell(row=3, column=2, value="All Amount in INR in Lakhs unless otherwise stated").font = font(size=9, italic=True, color="808080")

    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Particulars").font = font(bold=True)
    for y in range(1, layout.n_years+1):
        col = layout.bs_col(y)
        c = ws.cell(row=4, column=col, value=f"[Proj.]\nYear {y}")
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")

    def lbl(row, text, bold=False, color="000000", indent=0):
        ws.cell(row=row, column=2, value=("  "*indent)+text).font = font(bold=bold, color=color)

    def sec(row, text):
        ws.row_dimensions[row].height = 16
        c = ws.cell(row=row, column=2, value=text)
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+layout.n_years+1)

    def fr(row, fn, bold=False, bg=None, fmt=FMT_LAKHS):
        for y in range(1, layout.n_years+1):
            col = layout.bs_col(y); col_l = layout.bs_col_letter(y)
            c = ws.cell(row=row, column=col, value=fn(y, col_l))
            c.font = font(bold=bold); c.number_format = fmt
            if bg: c.fill = fill(bg)

    def tot(row, label, fn, fgcol="FFFFFF", bgcol="1F3864"):
        ws.row_dimensions[row].height = 17
        lbl(row, label, bold=True, color=fgcol)
        for y in range(1, layout.n_years+1):
            col = layout.bs_col(y); col_l = layout.bs_col_letter(y)
            c = ws.cell(row=row, column=col, value=fn(y, col_l))
            c.font = font(bold=True, color=fgcol)
            c.fill = fill(bgcol); c.number_format = FMT_LAKHS

    EXP = LayoutEngine
    tl_col  = lambda y: get_column_letter(EXP.TL_ANNUAL_COL_Y1+y-1)
    wc_col  = lambda y: get_column_letter(layout.wcap_col(y))
    cfs_col = lambda y: get_column_letter(layout.cfs_col(y))

    # Assumption refs for non-current items
    intangible_ref = layout.asmp_ref("balance.intangible_assets")
    nc_inv_ref     = layout.asmp_ref("balance.nc_investments")
    sec_dep_ref    = layout.asmp_ref("balance.security_deposits")
    other_nc_ref   = layout.asmp_ref("wc.other_non_current")
    fd_ref         = layout.asmp_ref("wc.investment_deposits")
    veh_loan_ref   = layout.asmp_ref("finance.vehicle_loan_amount")
    unsec_ref      = layout.asmp_ref("finance.unsecured_loan_amount")
    other_tl_ref   = layout.asmp_ref("finance.other_term_liab")

    # Cost & Means for total capex (gross block)
    from excel.sheet_costmeans import ASSET_DATA_START
    n_assets = len(store.capital_means.assets)
    cm_total_row = ASSET_DATA_START + n_assets + 1

    # ══════════════════════════════════════════════════════════════════════════
    # LIABILITIES
    # ══════════════════════════════════════════════════════════════════════════
    sec(5, "OUTSIDE LIABILITIES")

    lbl(layout.BS_TL_ROW,              "Term Loan",               indent=1)
    lbl(layout.BS_VEHICLE_LOAN_ROW,    "Vehicle Loan",            indent=1)
    lbl(layout.BS_UNSECURED_LOANS_ROW, "Unsecured Loans",         indent=1)
    lbl(layout.BS_OTHER_TERM_LIAB_ROW, "Other Term Liabilities",  indent=1)

    # TL outstanding from schedule
    fr(layout.BS_TL_ROW, lambda y,c: f"='Term Loan'!{tl_col(y)}{EXP.TL_ANNUAL_CLOSING_ROW}")
    # Vehicle loan: assumption amount, treated as constant (no amortization schedule)
    # For a complete model each loan needs its own schedule; here we show full amount declining linearly
    tl_tenor = store.capital_means.term_loans[0].tenor_months if store.capital_means.term_loans else 84
    fr(layout.BS_VEHICLE_LOAN_ROW, lambda y,c: f"=MAX({veh_loan_ref}*(1-{y}/{tl_tenor/12:.2f}),0)")
    # Unsecured loans: constant (bullet)
    fr(layout.BS_UNSECURED_LOANS_ROW, lambda y,c: f"={unsec_ref}")
    # Other term liabilities: constant
    fr(layout.BS_OTHER_TERM_LIAB_ROW, lambda y,c: f"={other_tl_ref}")

    tot(layout.BS_TOTAL_TERM_LIAB_ROW, "TOTAL TERM LIABILITIES",
        lambda y,c: (f"={c}{layout.BS_TL_ROW}+{c}{layout.BS_VEHICLE_LOAN_ROW}"
                     f"+{c}{layout.BS_UNSECURED_LOANS_ROW}+{c}{layout.BS_OTHER_TERM_LIAB_ROW}"),
        fgcol="FFFFFF", bgcol="2E75B6")

    lbl(layout.BS_TRADE_CRED_ROW, "Trade Creditors (WC)",  indent=1)
    lbl(layout.BS_OD_LOAN_ROW,    "WC Loan / OD Utilised", indent=1)
    fr(layout.BS_TRADE_CRED_ROW, lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_TOTAL_CL}")
    fr(layout.BS_OD_LOAN_ROW,    lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_WC_LOAN}")

    tot(layout.BS_TOTAL_CL_ROW, "TOTAL CURRENT LIABILITIES",
        lambda y,c: f"={c}{layout.BS_TRADE_CRED_ROW}+{c}{layout.BS_OD_LOAN_ROW}",
        fgcol="FFFFFF", bgcol="2E75B6")

    tot(layout.BS_TOTAL_OUTSIDE_LIAB, "TOTAL OUTSIDE LIABILITIES",
        lambda y,c: f"={c}{layout.BS_TOTAL_TERM_LIAB_ROW}+{c}{layout.BS_TOTAL_CL_ROW}")

    # Shareholders' Fund
    sec(layout.BS_EQUITY_ROW-1, "SHAREHOLDER'S FUND")
    # Share Capital = Promoter equity contribution + any additional from Assumption
    promoter_equity = store.capital_means.promoter_equity
    share_cap_add_ref = layout.asmp_ref("balance.share_capital_add")
    lbl(layout.BS_EQUITY_ROW,   "Share Capital",         indent=1)
    lbl(layout.BS_RESERVES_ROW, "Reserves & Surplus",    indent=1)

    fr(layout.BS_EQUITY_ROW, lambda y,c: f"={promoter_equity}+{share_cap_add_ref}")

    def reserves(y, c):
        terms = "+".join(f"PL!{layout.pl_col_letter(i)}{layout.PL_RETAINED_ROW}" for i in range(1, y+1))
        return f"={terms}"
    fr(layout.BS_RESERVES_ROW, reserves)

    tot(layout.BS_TOTAL_EQUITY_ROW, "Shareholder's Fund",
        lambda y,c: f"={c}{layout.BS_EQUITY_ROW}+{c}{layout.BS_RESERVES_ROW}",
        fgcol="FFFFFF", bgcol="2E75B6")

    tot(layout.BS_TOTAL_LIAB_ROW, "TOTAL LIABILITIES  (18-24)",
        lambda y,c: f"={c}{layout.BS_TOTAL_OUTSIDE_LIAB}+{c}{layout.BS_TOTAL_EQUITY_ROW}")

    # ══════════════════════════════════════════════════════════════════════════
    # ASSETS
    # ══════════════════════════════════════════════════════════════════════════
    sec(layout.BS_CASH_ROW-1, "ASSETS")

    # CURRENT ASSETS
    lbl(layout.BS_CASH_ROW,       "CURRENT ASSETS",             bold=True, color="1F3864")
    lbl(layout.BS_CASH_ROW,       "Cash & Bank Balance",        indent=1)
    lbl(layout.BS_FD_ROW,         "Fixed Deposits with Banks",  indent=1)
    lbl(layout.BS_DEBTORS_ROW,    "Receivables",                indent=1)
    lbl(layout.BS_DEBTORS_ROW+1,  "Inventory:",                 indent=1)
    lbl(layout.BS_CONSUMABLES_ROW,"  Consumables (RM Stock)",   indent=2)
    lbl(layout.BS_WIP_ROW,        "  Work in Progress",         indent=2)
    lbl(layout.BS_FG_ROW,         "  Finished Goods",           indent=2)
    lbl(layout.BS_COLD_STORE_ROW, "  Cold Store & Other Stores",indent=2)

    fr(layout.BS_CASH_ROW,        lambda y,c: f"=CFS!{cfs_col(y)}{layout.CFS_CLOSING_CASH}")
    fr(layout.BS_FD_ROW,          lambda y,c: f"={fd_ref}")
    fr(layout.BS_DEBTORS_ROW,     lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_DEBTORS}")
    fr(layout.BS_CONSUMABLES_ROW, lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_STOCK_RM}")
    fr(layout.BS_WIP_ROW,         lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_STOCK_WIP}")
    fr(layout.BS_FG_ROW,          lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_STOCK_FG}")
    fr(layout.BS_COLD_STORE_ROW,  lambda y,c: f"='W Cap'!{wc_col(y)}{layout.WCAP_STOCK_COLD}")

    tot(layout.BS_TOTAL_CA_ROW, "Total Current Assets",
        lambda y,c: (f"={c}{layout.BS_CASH_ROW}+{c}{layout.BS_FD_ROW}+{c}{layout.BS_DEBTORS_ROW}"
                     f"+{c}{layout.BS_CONSUMABLES_ROW}+{c}{layout.BS_WIP_ROW}"
                     f"+{c}{layout.BS_FG_ROW}+{c}{layout.BS_COLD_STORE_ROW}"),
        fgcol="1F3864", bgcol="EBF5FB")

    # FIXED ASSETS
    lbl(layout.BS_GROSS_BLOCK_ROW-1, "Fixed Assets", bold=True, color="1F3864")
    lbl(layout.BS_GROSS_BLOCK_ROW,   "Gross Block (WDV)")
    lbl(layout.BS_CUM_DEPR_ROW,      "Less: Cumulative Depreciation")
    lbl(layout.BS_NET_BLOCK_ROW,     "Net Block", bold=True)

    fr(layout.BS_GROSS_BLOCK_ROW, lambda y,c: f"='Cost & Means'!$D${cm_total_row}")
    total_depr_row = layout.dep_total_depr_row()
    def cum_depr(y, c):
        cols = "+".join(
            f"Depreciation!{get_column_letter(layout.DEP_COL_YEAR1+i-1)}{total_depr_row}"
            for i in range(1, y+1)
        )
        return f"={cols}"
    fr(layout.BS_CUM_DEPR_ROW, cum_depr)
    fr(layout.BS_NET_BLOCK_ROW,
       lambda y,c: f"={c}{layout.BS_GROSS_BLOCK_ROW}-{c}{layout.BS_CUM_DEPR_ROW}", bold=True)

    # NON-CURRENT ASSETS — from Assumption (user-input, not hardcoded zero)
    lbl(layout.BS_INTANGIBLE_ROW,    "Intangible Assets",         indent=1)
    lbl(layout.BS_NON_CURR_INV_ROW,  "Non Current Investments",   indent=1)
    lbl(layout.BS_SECURITY_DEP_ROW,  "Security Deposits",         indent=1)
    lbl(layout.BS_OTHER_NC_ROW,      "Other Non Current Assets",  indent=1)

    # These are Year 1 values from Assumption; held constant (no amortization)
    fr(layout.BS_INTANGIBLE_ROW,   lambda y,c: f"={intangible_ref}")
    fr(layout.BS_NON_CURR_INV_ROW, lambda y,c: f"={nc_inv_ref}")
    fr(layout.BS_SECURITY_DEP_ROW, lambda y,c: f"={sec_dep_ref}")
    fr(layout.BS_OTHER_NC_ROW,     lambda y,c: f"={other_nc_ref}")

    # TOTAL ASSETS
    tot(layout.BS_TOTAL_ASSETS_ROW, "TOTAL ASSETS  (34+35+39)",
        lambda y,c: (f"={c}{layout.BS_TOTAL_CA_ROW}+{c}{layout.BS_NET_BLOCK_ROW}"
                     f"+{c}{layout.BS_INTANGIBLE_ROW}+{c}{layout.BS_NON_CURR_INV_ROW}"
                     f"+{c}{layout.BS_SECURITY_DEP_ROW}+{c}{layout.BS_OTHER_NC_ROW}"))

    # BALANCE CHECK
    lbl(layout.BS_BALANCE_CHECK_ROW, "Balance Check  (Assets – Liabilities)", bold=True)
    for y in range(1, layout.n_years+1):
        col = layout.bs_col(y); col_l = layout.bs_col_letter(y)
        ok = chr(10003)+" BALANCED"; bad = chr(10007)+" GAP="
        c = ws.cell(row=layout.BS_BALANCE_CHECK_ROW, column=col,
                    value=(f'=IF(ROUND({col_l}{layout.BS_TOTAL_ASSETS_ROW},0)=ROUND({col_l}{layout.BS_TOTAL_LIAB_ROW},0),'
                           f'"{ok}","{bad}"&TEXT({col_l}{layout.BS_TOTAL_ASSETS_ROW}-{col_l}{layout.BS_TOTAL_LIAB_ROW},"#,##0.00"))'))
        c.font = font(bold=True, color="006400")

    # ── Bottom Ratios ────────────────────────────────────────────────────────
    sec(layout.BS_TNW_ROW-1, "KEY INDICATORS")
    lbl(layout.BS_TNW_ROW,              "TANGIBLE NET-WORTH",           bold=True)
    lbl(layout.BS_NWC_ROW,             "NET WORKING CAPITAL",           bold=True)
    lbl(layout.BS_CURR_RATIO_ROW,       "CURRENT RATIO",                bold=True)
    lbl(layout.BS_CURR_RATIO_NO_OD_ROW,"CURRENT RATIO  (Without OD)",  bold=True)
    lbl(layout.BS_TOL_TNW_ROW,          "TOL / TNW",                    bold=True)

    fr(layout.BS_TNW_ROW,
       lambda y,c: f"={c}{layout.BS_TOTAL_EQUITY_ROW}-{c}{layout.BS_INTANGIBLE_ROW}",
       bold=True, bg="F2F2F2")
    fr(layout.BS_NWC_ROW,
       lambda y,c: f"={c}{layout.BS_TOTAL_CA_ROW}-{c}{layout.BS_TOTAL_CL_ROW}",
       bold=True, bg="F2F2F2")
    fr(layout.BS_CURR_RATIO_ROW,
       lambda y,c: f"=IFERROR({c}{layout.BS_TOTAL_CA_ROW}/{c}{layout.BS_TOTAL_CL_ROW},0)",
       fmt="0.00", bold=True, bg="F2F2F2")
    fr(layout.BS_CURR_RATIO_NO_OD_ROW,
       lambda y,c: f"=IFERROR({c}{layout.BS_TOTAL_CA_ROW}/MAX({c}{layout.BS_TOTAL_CL_ROW}-{c}{layout.BS_OD_LOAN_ROW},1),0)",
       fmt="0.00", bold=True, bg="F2F2F2")
    fr(layout.BS_TOL_TNW_ROW,
       lambda y,c: f"=IFERROR({c}{layout.BS_TOTAL_OUTSIDE_LIAB}/MAX({c}{layout.BS_TNW_ROW},1),0)",
       fmt="0.00", bold=True, bg="F2F2F2")

    freeze_header(ws, row=5, col=layout.BS_COL_YEAR1)
    return ws
