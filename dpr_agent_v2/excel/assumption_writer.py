"""
excel/assumption_writer.py
═══════════════════════════
Writes the Assumption sheet — the single source of truth for the entire
DPR financial model.

Every cell address is computed via LayoutEngine.asmp_row().
Zero hardcoded row numbers in this file.

Color coding (industry standard):
  Blue  (#0070C0) — T1 user-provided input cells
  Green (#00B050) — T2 benchmark-filled cells
  Grey  (#808080) — T3 statutory/auto cells
  Black (#000000) — labels and headers
"""

from __future__ import annotations
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore, EntityType, AssetCategory
from core.layout_engine import LayoutEngine, COL_ID, COL_LABEL, COL_UNIT, COL_VALUE, COL_AUX
from core.assumption_registry import (
    Section, FieldTier, SECTION_LABEL,
    ANCHOR_A, ANCHOR_B, ANCHOR_C, ANCHOR_D, ANCHOR_E,
    ANCHOR_F, ANCHOR_G, ANCHOR_H, ANCHOR_I, ANCHOR_J,
)

# ─── Styles ───────────────────────────────────────────────────────────────────

FONT_NAME = "Arial"

def _font(bold=False, size=10, color="000000"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def _fill(hex_color: str):
    return PatternFill("solid", fgColor=hex_color)

def _border_bottom():
    return Border(bottom=Side(style="thin", color="CCCCCC"))

# Cell type fills
FILL_T1     = _fill("DDEEFF")   # light blue — user input
FILL_T2     = _fill("DDFFDD")   # light green — benchmark
FILL_T3     = _fill("F2F2F2")   # light grey — statutory
FILL_HEADER = _fill("1F3864")   # dark navy — section headers
FILL_SUB    = _fill("D6DCE4")   # light grey-blue — sub-headers

FONT_T1     = _font(color="00008B", size=10)    # dark blue text
FONT_T2     = _font(color="006400", size=10)    # dark green text
FONT_T3     = _font(color="404040", size=10)    # dark grey text
FONT_HEADER = _font(bold=True, color="FFFFFF", size=10)
FONT_LABEL  = _font(size=10)
FONT_ID     = _font(color="808080", size=9)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")


def _write_section_header(ws: Worksheet, row: int, label: str):
    """Write a full-width dark navy section header row."""
    ws.row_dimensions[row].height = 18
    cell = ws.cell(row=row, column=1, value=label)
    cell.font  = FONT_HEADER
    cell.fill  = FILL_HEADER
    cell.alignment = ALIGN_LEFT
    # Merge across columns A–F
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)


def _write_field_row(
    ws: Worksheet,
    row: int,
    field_id: str,
    label: str,
    unit: str,
    value,
    tier: FieldTier,
    aux_value=None,
    aux_label: str = "",
):
    """
    Write one data row in the Assumption sheet.
    Columns: A=ID, B=Label, C=spacer, D=Unit, E=Value (color coded), F=Aux
    """
    ws.row_dimensions[row].height = 16

    # Column A — field ID
    cell_id = ws.cell(row=row, column=COL_ID, value=field_id)
    cell_id.font      = FONT_ID
    cell_id.alignment = ALIGN_CENTER

    # Column B — label
    cell_label = ws.cell(row=row, column=COL_LABEL, value=label)
    cell_label.font      = FONT_LABEL
    cell_label.alignment = ALIGN_LEFT

    # Column D — unit
    cell_unit = ws.cell(row=row, column=COL_UNIT, value=unit)
    cell_unit.font      = FONT_T3
    cell_unit.alignment = ALIGN_RIGHT

    # Column E — value (color coded by tier)
    cell_val = ws.cell(row=row, column=COL_VALUE, value=value)
    cell_val.alignment = ALIGN_CENTER
    if tier == FieldTier.USER:
        cell_val.font = FONT_T1
        cell_val.fill = FILL_T1
    elif tier == FieldTier.BENCHMARK:
        cell_val.font = FONT_T2
        cell_val.fill = FILL_T2
    else:
        cell_val.font = FONT_T3
        cell_val.fill = FILL_T3

    # Format numbers
    if isinstance(value, float):
        if unit in ("fraction", "fraction p.a.", "fraction (0-1)") or "%" in unit:
            cell_val.number_format = '0.00%'
        elif "Lakhs" in unit or "INR" in unit:
            cell_val.number_format = '#,##0.00'
        elif unit in ("days", "months"):
            cell_val.number_format = '0'

    # Column F — auxiliary value (split%, input_per_output, etc.)
    if aux_value is not None:
        cell_aux = ws.cell(row=row, column=COL_AUX, value=aux_value)
        cell_aux.font      = FONT_T1 if tier == FieldTier.USER else FONT_T3
        cell_aux.fill      = FILL_T1 if tier == FieldTier.USER else FILL_T3
        cell_aux.alignment = ALIGN_CENTER
        if isinstance(aux_value, float) and 0 < aux_value <= 1.0 and "split" in aux_label.lower():
            cell_aux.number_format = '0.00%'


def write_assumption_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    """
    Write the complete Assumption sheet to the workbook.

    This is the only function that should write to the Assumption sheet.
    All values come from SessionStore. All row positions come from LayoutEngine.
    """
    ws = wb.create_sheet("Assumption", 0)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8   # ID
    ws.column_dimensions["B"].width = 45  # Label
    ws.column_dimensions["C"].width = 3   # Spacer
    ws.column_dimensions["D"].width = 22  # Unit
    ws.column_dimensions["E"].width = 14  # Value
    ws.column_dimensions["F"].width = 14  # Aux value

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    title_cell = ws.cell(row=1, column=1,
                         value=f"ASSUMPTIONS & INPUTS  —  {store.project_profile.company_name}")
    title_cell.font = _font(bold=True, size=12, color="1F3864")
    ws.merge_cells("A1:F1")

    ws.cell(row=2, column=1,
            value="(All monetary values in INR Lakhs unless otherwise stated)")
    ws.cell(row=2, column=1).font = _font(size=9, color="808080")
    ws.merge_cells("A2:F2")

    # ── Section A: Capacity ───────────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_A - 1, SECTION_LABEL[Section.A])

    _write_field_row(ws, layout.asmp_row("capacity.year1_util"),
                     "A1", "Year 1 Capacity Utilisation",
                     "fraction (0-1)", store.revenue_model.year1_utilization,
                     FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("capacity.annual_increment"),
                     "A2", "Annual Utilisation Increment",
                     "fraction p.a.", store.revenue_model.annual_utilization_increment,
                     FieldTier.BENCHMARK)
    _write_field_row(ws, layout.asmp_row("capacity.max_util"),
                     "A3", "Maximum Utilisation Ceiling",
                     "fraction (0-1)", store.revenue_model.max_utilization,
                     FieldTier.BENCHMARK)
    _write_field_row(ws, layout.asmp_row("capacity.working_days"),
                     "A4", "Working Days per Month",
                     "days", store.revenue_model.working_days_per_month,
                     FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("capacity.months_in_year"),
                     "A5", "Months in a Year",
                     "months", 12,
                     FieldTier.STATUTORY)

    # ── Section B: Revenue — one block per product ────────────────────────────
    _write_section_header(ws, ANCHOR_B - 1, SECTION_LABEL[Section.B])

    # Column headers for this section
    ws.cell(row=ANCHOR_B, column=COL_LABEL, value="Product / Service Name").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_B, column=COL_UNIT,  value="Unit").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_B, column=COL_VALUE, value="Base Price (Yr 1)").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_B, column=COL_AUX,   value="Capacity / Split %").font = _font(bold=True, size=9)

    for i, product in enumerate(store.revenue_model.products):
        prod_id = f"B{i+1}"

        # Row 0: product name + unit + output_ratio (col F)
        name_row = layout.asmp_row("revenue.product.name", i)
        ws.row_dimensions[name_row].height = 16
        ws.cell(row=name_row, column=COL_ID,    value=prod_id).font = FONT_ID
        ws.cell(row=name_row, column=COL_LABEL, value=product.name).font = _font(bold=True)
        ws.cell(row=name_row, column=COL_UNIT,  value=product.unit).font = FONT_T1
        ws.cell(row=name_row, column=COL_VALUE, value="").fill = FILL_T1
        # Output ratio in col F — used by Revenue sheet formulas
        out_cell = ws.cell(row=name_row, column=COL_AUX, value=product.output_ratio)
        out_cell.font      = FONT_T1
        out_cell.fill      = FILL_T1
        out_cell.alignment = ALIGN_CENTER
        out_cell.number_format = "0.000"

        # Row 1: base price
        price_row = layout.asmp_row("revenue.product.price", i)
        _write_field_row(ws, price_row,
                         "", f"  Base Selling Price (Year 1)",
                         f"₹/{product.unit}", product.price_per_unit,
                         FieldTier.USER)

        # Row 2: capacity per day + split %
        cap_row = layout.asmp_row("revenue.product.capacity", i)
        _write_field_row(ws, cap_row,
                         "", f"  Capacity per Day",
                         f"{product.unit}/day", product.capacity_per_day,
                         FieldTier.USER,
                         aux_value=product.split_percent,
                         aux_label="split_%")

        # Row 3: price escalation
        esc_row = layout.asmp_row("revenue.product.escalation", i)
        _write_field_row(ws, esc_row,
                         "", f"  Annual Price Escalation",
                         "fraction p.a.", product.price_escalation,
                         FieldTier.BENCHMARK)

    # ── Section C: Raw Materials — one block per material ─────────────────────
    _write_section_header(ws, ANCHOR_C - 1, SECTION_LABEL[Section.C])

    ws.cell(row=ANCHOR_C, column=COL_LABEL, value="Material Name").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_C, column=COL_UNIT,  value="Unit").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_C, column=COL_VALUE, value="Base Price (Yr 1)").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_C, column=COL_AUX,   value="Input per Output Unit").font = _font(bold=True, size=9)

    for i, mat in enumerate(store.cost_structure.raw_materials):
        mat_id = f"C{i+1}"

        # Row 0: material name + unit
        name_row = layout.asmp_row("material.name", i)
        ws.row_dimensions[name_row].height = 16
        ws.cell(row=name_row, column=COL_ID,    value=mat_id).font = FONT_ID
        ws.cell(row=name_row, column=COL_LABEL, value=mat.name).font = _font(bold=True)
        ws.cell(row=name_row, column=COL_UNIT,  value=f"per {mat.unit}").font = FONT_T1
        ws.cell(row=name_row, column=COL_VALUE, value="").fill = FILL_T1

        # Row 1: base price + input_per_output in aux column
        price_row = layout.asmp_row("material.price", i)
        _write_field_row(ws, price_row,
                         "", f"  Base Price per {mat.unit}",
                         f"₹/{mat.unit}", mat.price_per_unit,
                         FieldTier.USER,
                         aux_value=mat.input_per_output,
                         aux_label="input_per_output")
        # Label the aux col header for first material
        if i == 0:
            ws.cell(row=price_row, column=COL_AUX,
                    value=mat.input_per_output).number_format = "0.000"

        # Row 2: escalation
        esc_row = layout.asmp_row("material.escalation", i)
        _write_field_row(ws, esc_row,
                         "", "  Annual Cost Escalation",
                         "fraction p.a.", mat.price_escalation,
                         FieldTier.BENCHMARK)

    # ── Section D: Operating Expenses ────────────────────────────────────────
    _write_section_header(ws, ANCHOR_D - 1, SECTION_LABEL[Section.D])

    cs = store.cost_structure
    opex_fields = [
        ("opex.rm_pct_fa",          "D1",  "Repair & Maintenance",         "% of Net Fixed Assets",  cs.rm_pct_of_fa,          FieldTier.BENCHMARK, None),
        ("opex.rm_escalation",      "D1e", "  └ R&M Escalation Rate",      "fraction p.a.",           cs.rm_escalation,         FieldTier.BENCHMARK, None),
        ("opex.insurance_pct_fa",   "D2",  "Insurance",                    "% of Net Fixed Assets",  cs.insurance_pct_of_fa,   FieldTier.BENCHMARK, None),
        ("opex.insurance_escalation","D2e","  └ Insurance Escalation Rate","fraction p.a.",           cs.insurance_escalation,  FieldTier.BENCHMARK, None),
        ("opex.power_pct_revenue",  "D3",  "Power & Fuel",                 "% of Revenue",            cs.power_pct_revenue,     FieldTier.BENCHMARK, None),
        ("opex.power_escalation",   "D3e", "  └ Power Cost Escalation",    "fraction p.a.",           cs.power_escalation,      FieldTier.BENCHMARK, None),
        ("opex.marketing_pct_revenue","D4","Marketing Expenses",           "% of Revenue",            cs.marketing_pct_revenue, FieldTier.BENCHMARK, None),
        ("opex.marketing_escalation","D4e","  └ Marketing Escalation",     "fraction p.a.",           cs.marketing_escalation,  FieldTier.STATUTORY, None),
        ("opex.transport_base",     "D5",  "Transportation Cost (Base)",   "INR Lakhs (Year 1)",      cs.transport_base_lakhs,  FieldTier.USER,       None),
        ("opex.transport_escalation","D5e","  └ Transport Escalation",     "fraction p.a.",           cs.transport_escalation,  FieldTier.BENCHMARK,  None),
        ("opex.misc_base",          "D6",  "Miscellaneous Expenses (Base)","INR Lakhs (Year 1)",      cs.misc_base_lakhs,       FieldTier.USER,       None),
        ("opex.misc_escalation",    "D6e", "  └ Misc Escalation",          "fraction p.a.",           cs.misc_escalation,       FieldTier.BENCHMARK,  None),
        ("opex.sga_base",           "D7",  "Selling, General & Admin (Base)","INR Lakhs (Year 1)",   cs.sga_base_lakhs,        FieldTier.BENCHMARK,  None),
        ("opex.sga_escalation",     "D7e", "  └ SGA Escalation",           "fraction p.a.",           cs.sga_escalation,        FieldTier.BENCHMARK,  None),
    ]

    for key, fid, label, unit, value, tier, aux in opex_fields:
        row = layout.asmp_row(key)
        _write_field_row(ws, row, fid, label, unit, value, tier, aux)

    # ── Section E: Manpower ───────────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_E - 1, SECTION_LABEL[Section.E])

    ws.cell(row=ANCHOR_E, column=COL_LABEL, value="Designation").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_E, column=COL_UNIT,  value="Head Count").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_E, column=COL_VALUE, value="Monthly Salary (Lakhs)").font = _font(bold=True, size=9)
    ws.cell(row=ANCHOR_E, column=COL_AUX,   value="Annual Increment").font = _font(bold=True, size=9)

    for i, emp in enumerate(store.manpower.categories):
        emp_id = f"E{i+1}"

        # Row 0: designation + headcount in unit column (reused)
        desig_row = layout.asmp_row("employee.designation", i)
        ws.row_dimensions[desig_row].height = 16
        ws.cell(row=desig_row, column=COL_ID,    value=emp_id).font = FONT_ID
        ws.cell(row=desig_row, column=COL_LABEL, value=emp.designation).font = _font(bold=True)
        # Headcount goes in unit column (D), styled as T1
        hc_cell = ws.cell(row=desig_row, column=COL_UNIT, value=emp.count)
        hc_cell.font = FONT_T1
        hc_cell.fill = FILL_T1
        hc_cell.alignment = ALIGN_CENTER

        # Row 1: monthly salary + annual increment
        sal_row = layout.asmp_row("employee.salary", i)
        _write_field_row(ws, sal_row,
                         "", "  Monthly Salary",
                         "INR Lakhs/month", emp.monthly_salary_lakhs,
                         FieldTier.USER,
                         aux_value=emp.annual_increment,
                         aux_label="increment")

    # ── Section F: Finance ────────────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_F - 1, SECTION_LABEL[Section.F])

    cm = store.capital_means
    tls = cm.term_loans
    ods = cm.od_sources

    if tls:
        tl = tls[0]
        repayment_months = tl.tenor_months - tl.moratorium_months
        ws.cell(row=ANCHOR_F, column=COL_LABEL,
                value=f"  Term Loan — {tl.label}").font = _font(bold=True, size=9, color="1F3864")

        _write_field_row(ws, layout.asmp_row("finance.tl_amount"),
                         "F1a", "Loan Amount", "INR Lakhs", tl.amount_lakhs,
                         FieldTier.USER)
        _write_field_row(ws, layout.asmp_row("finance.tl_rate"),
                         "F1b", "Annual Interest Rate", "fraction p.a.", tl.interest_rate,
                         FieldTier.USER)
        _write_field_row(ws, layout.asmp_row("finance.tl_tenor"),
                         "F1c", "Total Tenor", "months", tl.tenor_months,
                         FieldTier.USER)
        _write_field_row(ws, layout.asmp_row("finance.tl_moratorium"),
                         "F1d", "Moratorium Period", "months", tl.moratorium_months,
                         FieldTier.USER)

        # Repayment months = derived formula
        rep_row = layout.asmp_row("finance.tl_repayment")
        rep_cell = ws.cell(row=rep_row, column=COL_VALUE,
                           value=f"={layout.asmp_addr('finance.tl_tenor')}-{layout.asmp_addr('finance.tl_moratorium')}")
        rep_cell.font = FONT_T3
        rep_cell.fill = FILL_T3
        ws.cell(row=rep_row, column=COL_LABEL, value="  Repayment Months (derived)").font = FONT_LABEL
        ws.cell(row=rep_row, column=COL_UNIT,  value="months").font = FONT_T3

    if ods:
        od = ods[0]
        _write_field_row(ws, layout.asmp_row("finance.od_limit"),
                         "F_OD1", "OD / CC Limit", "INR Lakhs", od.amount_lakhs,
                         FieldTier.USER)
        _write_field_row(ws, layout.asmp_row("finance.od_rate"),
                         "F_OD2", "OD Interest Rate", "fraction p.a.", od.interest_rate,
                         FieldTier.BENCHMARK)

    # ── Section G: Working Capital ────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_G - 1, SECTION_LABEL[Section.G])

    wc = store.working_capital
    od_limit = ods[0].amount_lakhs if ods else 0.0
    od_rate  = ods[0].interest_rate if ods else 0.0

    wc_fields = [
        ("wc.debtor_days",         "G1", "Debtor Days (Receivables)",      "days", wc.debtor_days,         FieldTier.USER),
        ("wc.creditor_days_rm",    "G2", "Creditor Days — Raw Materials",   "days", wc.creditor_days_rm,    FieldTier.USER),
        ("wc.creditor_days_admin", "G3", "Creditor Days — Admin Expenses",  "days", wc.creditor_days_admin, FieldTier.BENCHMARK),
        ("wc.stock_days_rm",       "G4", "Raw Material Stock Days",         "days", wc.stock_days_rm,       FieldTier.USER),
        ("wc.stock_days_fg",       "G5", "Finished Goods Stock Days",       "days", wc.stock_days_fg,       FieldTier.BENCHMARK),
        ("wc.wc_loan_amount",      "G6", "Working Capital Loan",            "INR Lakhs", od_limit,          FieldTier.STATUTORY),
        ("wc.wc_interest_rate",    "G7", "Working Capital Interest Rate",   "fraction p.a.", od_rate,       FieldTier.STATUTORY),
    ]
    for key, fid, label, unit, value, tier in wc_fields:
        _write_field_row(ws, layout.asmp_row(key), fid, label, unit, value, tier)

    # ── Section H: Depreciation ───────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_H - 1, SECTION_LABEL[Section.H])

    dep_fields = [
        ("depr.plant_machinery", "H1", "Plant & Machinery",       0.15),
        ("depr.civil_works",     "H2", "Civil Works / Building",  0.10),
        ("depr.furniture",       "H3", "Furniture & Fixtures",    0.10),
        ("depr.vehicle",         "H4", "Vehicles",                0.15),
        ("depr.electrical",      "H5", "Electrical & Fittings",   0.10),
        ("depr.other",           "H6", "Other Assets",            0.15),
    ]
    for key, fid, label, rate in dep_fields:
        _write_field_row(ws, layout.asmp_row(key),
                         fid, label, "WDV % p.a.", rate, FieldTier.STATUTORY)

    # ── Section I: Implementation ─────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_I - 1, SECTION_LABEL[Section.I])

    _write_field_row(ws, layout.asmp_row("impl.months"),
                     "I1", "Implementation Period (months before COD)",
                     "months", store.working_capital.implementation_months,
                     FieldTier.USER)

    # ── Section J: Tax ────────────────────────────────────────────────────────
    _write_section_header(ws, ANCHOR_J - 1, SECTION_LABEL[Section.J])

    _write_field_row(ws, layout.asmp_row("tax.entity_type"),
                     "J1", "Entity Type", "type",
                     store.project_profile.entity_type.value, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("tax.company_rate"),
                     "J2", "Company Tax Rate", "fraction", 0.25, FieldTier.STATUTORY)
    _write_field_row(ws, layout.asmp_row("tax.surcharge"),
                     "J3", "Surcharge", "fraction", 0.07, FieldTier.STATUTORY)
    _write_field_row(ws, layout.asmp_row("tax.hec"),
                     "J4", "Health & Education Cess", "fraction", 0.04, FieldTier.STATUTORY)
    _write_field_row(ws, layout.asmp_row("tax.llp_rate"),
                     "J5", "LLP / Proprietorship Tax Rate", "fraction", 0.30, FieldTier.STATUTORY)
    _write_field_row(ws, layout.asmp_row("tax.surcharge_threshold"),
                     "J6", "Surcharge Threshold", "INR Lakhs", 100.0, FieldTier.STATUTORY)


    # ── K: Balance Sheet Non-Current Items ──────────────────────────────────────
    from core.assumption_registry import ANCHOR_K
    ws.cell(row=ANCHOR_K - 1, column=1, value="K").font = _font(bold=True, color="FFFFFF")
    ws.cell(row=ANCHOR_K - 1, column=1).fill = _fill("1F3864")
    ws.cell(row=ANCHOR_K - 1, column=2, value="K  |  BALANCE SHEET — NON-CURRENT ITEMS").font = _font(bold=True)
    _write_field_row(ws, layout.asmp_row("balance.intangible_assets"),
                     "K1", "Intangible Assets (Year 1 Book Value)", "INR Lakhs", 0.0, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("balance.nc_investments"),
                     "K2", "Non-Current Investments", "INR Lakhs", 0.0, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("balance.security_deposits"),
                     "K3", "Security Deposits", "INR Lakhs", 0.0, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("balance.share_capital_add"),
                     "K4", "Additional Share Capital (beyond equity contribution)", "INR Lakhs", 0.0, FieldTier.USER)
    # Additional vehicle/unsecured/other term liab (written under Section F block)
    _write_field_row(ws, layout.asmp_row("finance.vehicle_loan_amount"),
                     "F9", "Vehicle Loan — Outstanding", "INR Lakhs", 0.0, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("finance.unsecured_loan_amount"),
                     "F10", "Unsecured Loans", "INR Lakhs", 0.0, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("finance.other_term_liab"),
                     "F11", "Other Term Liabilities", "INR Lakhs", 0.0, FieldTier.USER)
    # WIP & Cold Store days (written under Section G block)
    _write_field_row(ws, layout.asmp_row("wc.wip_days"),
                     "G10", "Work-in-Progress Days", "days", 0, FieldTier.USER)
    _write_field_row(ws, layout.asmp_row("wc.cold_store_days"),
                     "G11", "Cold Store / Other Store Days", "days", 0, FieldTier.USER)

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = ANCHOR_J + 8
    ws.cell(row=legend_row, column=1, value="LEGEND:").font = _font(bold=True, size=9)
    for col, (fill, label) in enumerate([
        (FILL_T1, "Blue = User provided (T1)"),
        (FILL_T2, "Green = Benchmarked (T2)"),
        (FILL_T3, "Grey = Statutory (T3)"),
    ], start=2):
        c = ws.cell(row=legend_row, column=col, value=label)
        c.fill = fill
        c.font = _font(size=9)

    return ws
