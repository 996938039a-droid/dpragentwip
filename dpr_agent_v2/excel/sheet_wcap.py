"""excel/sheet_wcap.py — Working Capital with RM, FG, WIP, Cold Store stock rows."""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL,
    FMT_LAKHS, font, fill
)

def write_wcap_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("W Cap")
    set_col_widths(ws, {
        "A": 4, "B": 40, "C": 16,
        **{get_column_letter(layout.wcap_col(y)): 13 for y in range(1, layout.n_years+1)}
    })
    write_title(ws, store.project_profile.company_name, "Working Capital Computation")

    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Particulars").font = font(bold=True)
    ws.cell(row=4, column=3, value="Basis").font = font(bold=True, size=9, color="808080")
    for y in range(1, layout.n_years+1):
        col = layout.wcap_col(y)
        c = ws.cell(row=4, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")

    def _yl(y): return get_column_letter(layout.wcap_col(y))

    # Assumption refs
    cr_rm  = layout.asmp_ref("wc.creditor_days_rm")
    cr_adm = layout.asmp_ref("wc.creditor_days_admin")
    st_rm  = layout.asmp_ref("wc.stock_days_rm")
    st_fg  = layout.asmp_ref("wc.stock_days_fg")
    st_wip = layout.asmp_ref("wc.wip_days")
    st_cold= layout.asmp_ref("wc.cold_store_days")
    deb    = layout.asmp_ref("wc.debtor_days")
    od_ref = layout.asmp_ref("wc.wc_loan_amount")
    rate   = layout.asmp_ref("wc.wc_interest_rate")

    def rev_ref(y):  return f"Revenue!{layout.year_col_letter(y)}{layout.rev_total_row()}"
    def cogs_ref(y): return f"Expenses!{layout.year_col_letter(y)}{layout.exp_total_cogs_row()}"

    # ── A: Current Liabilities ────────────────────────────────────────────────
    ws.cell(row=layout.WCAP_CREDITORS_RM-1, column=2, value="A  |  CURRENT LIABILITIES").font = font(bold=True, color="FFFFFF")
    ws.cell(row=layout.WCAP_CREDITORS_RM-1, column=2).fill = fill("2E75B6")

    ws.cell(row=layout.WCAP_CREDITORS_RM,    column=2, value="  Trade Creditors (Raw Materials)").font = FONT_BODY
    ws.cell(row=layout.WCAP_CREDITORS_ADMIN, column=2, value="  Creditors (Admin / Other Expenses)").font = FONT_BODY
    ws.cell(row=layout.WCAP_TOTAL_CL,        column=2, value="TOTAL CURRENT LIABILITIES").font = FONT_TOTAL
    ws.cell(row=layout.WCAP_CREDITORS_RM,    column=3, value="COGS/365 × days").font = font(size=9, color="808080")
    ws.cell(row=layout.WCAP_CREDITORS_ADMIN, column=3, value="Revenue/365 × days").font = font(size=9, color="808080")

    for y in range(1, layout.n_years+1):
        col = layout.wcap_col(y); col_l = _yl(y)
        ws.cell(row=layout.WCAP_CREDITORS_RM,    column=col, value=f"={cogs_ref(y)}/365*{cr_rm}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_CREDITORS_ADMIN, column=col, value=f"={rev_ref(y)}/365*{cr_adm}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_TOTAL_CL,        column=col,
                value=f"={col_l}{layout.WCAP_CREDITORS_RM}+{col_l}{layout.WCAP_CREDITORS_ADMIN}").font = FONT_TOTAL
        for r in [layout.WCAP_CREDITORS_RM, layout.WCAP_CREDITORS_ADMIN, layout.WCAP_TOTAL_CL]:
            ws.cell(row=r, column=col).number_format = FMT_LAKHS

    # ── B: Current Assets ─────────────────────────────────────────────────────
    ws.cell(row=layout.WCAP_STOCK_RM-1, column=2, value="B  |  CURRENT ASSETS").font = font(bold=True, color="FFFFFF")
    ws.cell(row=layout.WCAP_STOCK_RM-1, column=2).fill = fill("2E75B6")

    ws.cell(row=layout.WCAP_STOCK_RM,   column=2, value="  Raw Material Closing Stock").font = FONT_BODY
    ws.cell(row=layout.WCAP_STOCK_FG,   column=2, value="  Finished Goods Stock").font = FONT_BODY
    ws.cell(row=layout.WCAP_STOCK_WIP,  column=2, value="  Work-in-Progress (WIP)").font = FONT_BODY
    ws.cell(row=layout.WCAP_STOCK_COLD, column=2, value="  Cold Store & Other Stores").font = FONT_BODY
    ws.cell(row=layout.WCAP_DEBTORS,    column=2, value="  Trade Receivables (Debtors)").font = FONT_BODY
    ws.cell(row=layout.WCAP_TOTAL_CA,   column=2, value="TOTAL CURRENT ASSETS").font = FONT_TOTAL

    ws.cell(row=layout.WCAP_STOCK_RM,   column=3, value="COGS/365 × days").font = font(size=9, color="808080")
    ws.cell(row=layout.WCAP_STOCK_FG,   column=3, value="COGS/365 × FG days").font = font(size=9, color="808080")
    ws.cell(row=layout.WCAP_STOCK_WIP,  column=3, value="COGS/365 × WIP days").font = font(size=9, color="808080")
    ws.cell(row=layout.WCAP_STOCK_COLD, column=3, value="COGS/365 × cold days").font = font(size=9, color="808080")
    ws.cell(row=layout.WCAP_DEBTORS,    column=3, value="Revenue/365 × days").font = font(size=9, color="808080")

    for y in range(1, layout.n_years+1):
        col = layout.wcap_col(y); col_l = _yl(y)
        ws.cell(row=layout.WCAP_STOCK_RM,   column=col, value=f"={cogs_ref(y)}/365*{st_rm}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_STOCK_FG,   column=col, value=f"={cogs_ref(y)}/365*{st_fg}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_STOCK_WIP,  column=col, value=f"={cogs_ref(y)}/365*{st_wip}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_STOCK_COLD, column=col, value=f"={cogs_ref(y)}/365*{st_cold}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_DEBTORS,    column=col, value=f"={rev_ref(y)}/365*{deb}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_TOTAL_CA,   column=col,
                value=(f"={col_l}{layout.WCAP_STOCK_RM}+{col_l}{layout.WCAP_STOCK_FG}"
                       f"+{col_l}{layout.WCAP_STOCK_WIP}+{col_l}{layout.WCAP_STOCK_COLD}"
                       f"+{col_l}{layout.WCAP_DEBTORS}")).font = FONT_TOTAL
        for r in [layout.WCAP_STOCK_RM, layout.WCAP_STOCK_FG, layout.WCAP_STOCK_WIP,
                  layout.WCAP_STOCK_COLD, layout.WCAP_DEBTORS, layout.WCAP_TOTAL_CA]:
            ws.cell(row=r, column=col).number_format = FMT_LAKHS

    # ── WC Requirement ────────────────────────────────────────────────────────
    ws.cell(row=layout.WCAP_WC_REQUIREMENT, column=2, value="NET WORKING CAPITAL REQUIREMENT").font = FONT_TOTAL
    ws.cell(row=layout.WCAP_WC_LOAN,        column=2, value="  WC Loan / OD Limit (Sanctioned)").font = FONT_BODY
    ws.cell(row=layout.WCAP_WC_INTEREST,    column=2, value="  Working Capital Interest").font = FONT_BODY

    for y in range(1, layout.n_years+1):
        col = layout.wcap_col(y); col_l = _yl(y)
        ws.cell(row=layout.WCAP_WC_REQUIREMENT, column=col,
                value=f"={col_l}{layout.WCAP_TOTAL_CA}-{col_l}{layout.WCAP_TOTAL_CL}").font = FONT_TOTAL
        # OD Loan = directly from Assumption (sanctioned limit)
        ws.cell(row=layout.WCAP_WC_LOAN, column=col, value=f"={od_ref}").font = FONT_FORMULA
        ws.cell(row=layout.WCAP_WC_INTEREST, column=col,
                value=f"={col_l}{layout.WCAP_WC_LOAN}*{rate}").font = FONT_FORMULA
        for r in [layout.WCAP_WC_REQUIREMENT, layout.WCAP_WC_LOAN, layout.WCAP_WC_INTEREST]:
            ws.cell(row=r, column=col).number_format = FMT_LAKHS
    return ws
