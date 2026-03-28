"""
excel/sheet_manpower.py
═════════════════════════
Manpower sheet — computes annual salary costs per category
with annual increment escalation.
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_REF,
    FILL_TOTAL, AL_LEFT, AL_CENTER,
    FMT_LAKHS, FMT_INT,
    font, fill
)


def write_manpower_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("ManPower")

    set_col_widths(ws, {
        "A": 4, "B": 30, "C": 10, "D": 14, "E": 14, "F": 14,
        **{get_column_letter(layout.year_col(y)): 13
           for y in range(1, layout.n_years + 1)}
    })

    write_title(ws, store.project_profile.company_name, "Manpower Expenses")

    # ── Static header row ─────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    for col, label in [(2,"Designation"), (3,"Count"), (4,"Monthly ₹L"),
                       (5,"Annual ₹L"), (6,"Annual Total ₹L")]:
        c = ws.cell(row=4, column=col, value=label)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    # ── Employee rows ─────────────────────────────────────────────────────────
    for e_idx, emp in enumerate(store.manpower.categories):
        row = layout.mp_employee_row(e_idx)

        ws.cell(row=row, column=2, value=emp.designation).font = FONT_BODY
        # Count from Assumption (col D holds headcount in our layout)
        count_ref = layout.asmp_ref("employee.designation", e_idx, col=4)  # col D
        ws.cell(row=row, column=3,
                value=f"={layout.asmp_ref('employee.designation', e_idx, col=4)}").font = FONT_REF
        # Monthly salary from Assumption
        ws.cell(row=row, column=4,
                value=f"={layout.asmp_ref('employee.salary', e_idx)}").font = FONT_REF
        # Annual per person = monthly × 12
        ws.cell(row=row, column=5, value=f"=D{row}*12").font = FONT_FORMULA
        # Annual total = count × annual
        ws.cell(row=row, column=6, value=f"=C{row}*E{row}").font = FONT_FORMULA
        for col in [4, 5, 6]:
            ws.cell(row=row, column=col).number_format = FMT_LAKHS

    # ── Base year total ────────────────────────────────────────────────────────
    base_row = layout.mp_base_total_row()
    ws.cell(row=base_row, column=2, value="Base Year Annual Salary").font = font(bold=True)
    first_emp = layout.mp_employee_row(0)
    last_emp  = layout.mp_employee_row(layout.n_employees - 1)
    ws.cell(row=base_row, column=6,
            value=f"=SUM(F{first_emp}:F{last_emp})").font = FONT_TOTAL
    ws.cell(row=base_row, column=6).number_format = FMT_LAKHS

    # ── Annual projections ────────────────────────────────────────────────────
    ann_row = layout.mp_annual_row()
    ws.row_dimensions[ann_row - 1].height = 16
    ws.cell(row=ann_row - 1, column=2,
            value="Annual Salary Projections").font = font(bold=True, color="1F3864")

    # Year headers
    for y in range(1, layout.n_years + 1):
        col = layout.year_col(y)
        c = ws.cell(row=ann_row - 1, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    ws.cell(row=ann_row, column=2,
            value="Total Annual Salary (Fixed)").font = font(bold=True)
    ws.cell(row=ann_row, column=3, value="₹ Lakhs").font = font(size=9, color="808080")

    # Use the first employee's increment rate for all (they all have the same T2 benchmark)
    inc_ref = layout.asmp_ref("employee.increment", 0, col=6)  # col F = increment in our layout

    for y in range(1, layout.n_years + 1):
        col   = layout.year_col(y)
        col_l = layout.year_col_letter(y)
        if y == 1:
            ws.cell(row=ann_row, column=col,
                    value=f"=F{base_row}").font = FONT_REF
        else:
            prev_col = layout.year_col_letter(y - 1)
            ws.cell(row=ann_row, column=col,
                    value=f"={prev_col}{ann_row}*(1+{inc_ref})").font = FONT_FORMULA
        ws.cell(row=ann_row, column=col).number_format = FMT_LAKHS
        ws.cell(row=ann_row, column=col).fill = fill("EBF5FB")

    # Transfer to P&L row
    pl_row = ann_row + 2
    ws.cell(row=pl_row, column=2,
            value="Salary Expense (Annual, for P&L)").font = font(bold=True)
    ws.cell(row=pl_row, column=3, value="₹ Lakhs").font = font(size=9, color="808080")

    for y in range(1, layout.n_years + 1):
        col   = layout.year_col(y)
        col_l = layout.year_col_letter(y)
        ws.cell(row=pl_row, column=col,
                value=f"={col_l}{ann_row}").font = FONT_FORMULA
        ws.cell(row=pl_row, column=col).number_format = FMT_LAKHS

    freeze_header(ws, row=5, col=3)
    return ws
