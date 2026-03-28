"""excel/sheet_pl.py — Projected P&L with Drawings, ICR, EBITDA Margin."""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_REF,
    FMT_LAKHS, FMT_ZERO, FMT_PCT,
    font, fill
)

def write_pl_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("PL")
    set_col_widths(ws, {"A": 4, "B": 40, "C": 6,
                        **{get_column_letter(layout.pl_col(y)): 13 for y in range(1, layout.n_years+1)}})
    write_title(ws, store.project_profile.company_name, "Projected Profit & Loss Account")

    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Particulars").font = font(bold=True)
    ws.cell(row=4, column=3, value="Sch").font = font(bold=True, size=9, color="808080")
    for y in range(1, layout.n_years+1):
        col = layout.pl_col(y)
        c = ws.cell(row=4, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")

    def lbl(row, text, bold=False, color="000000", sch=""):
        ws.cell(row=row, column=2, value=text).font = font(bold=bold, color=color)
        if sch: ws.cell(row=row, column=3, value=sch).font = font(size=9, color="808080")

    def sec(row, text):
        ws.row_dimensions[row].height = 16
        c = ws.cell(row=row, column=2, value=text)
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3+layout.n_years)

    def fr(row, fn, fmt=FMT_LAKHS, bold=False, bg=None):
        for y in range(1, layout.n_years+1):
            col = layout.pl_col(y); col_l = layout.pl_col_letter(y)
            c = ws.cell(row=row, column=col, value=fn(y, col_l))
            c.font = font(bold=bold); c.number_format = fmt
            if bg: c.fill = fill(bg)

    pl = layout.pl_col_letter
    yl = layout.year_col_letter
    EXP = LayoutEngine

    # I. Revenue
    sec(layout.PL_REVENUE_ROW - 1, "I.  REVENUE FROM OPERATIONS")
    lbl(layout.PL_REVENUE_ROW, "Total Revenue from Operations", bold=True, sch="Revenue")
    fr(layout.PL_REVENUE_ROW, lambda y,c: f"=Revenue!{yl(y)}{layout.rev_total_row()}", bold=True)

    # II. COGS
    sec(layout.PL_COGS_ROW - 1, "II.  COST OF SALES")
    lbl(layout.PL_COGS_ROW, "  Raw Material Cost of Sales", sch="Expenses")
    fr(layout.PL_COGS_ROW, lambda y,c: f"=Expenses!{yl(y)}{layout.exp_total_cogs_row()}")
    lbl(layout.PL_COGS_ROW+1, "Total Cost of Sales", bold=True)
    fr(layout.PL_COGS_ROW+1, lambda y,c: f"={c}{layout.PL_COGS_ROW}", bold=True)

    # Gross Profit
    ws.row_dimensions[layout.PL_GROSS_PROFIT_ROW].height = 17
    lbl(layout.PL_GROSS_PROFIT_ROW, "GROSS PROFIT  (Revenue – COGS)", bold=True, color="1F3864")
    fr(layout.PL_GROSS_PROFIT_ROW, lambda y,c: f"={c}{layout.PL_REVENUE_ROW}-{c}{layout.PL_COGS_ROW+1}", bold=True, bg="EBF5FB")

    # III. Operating Expenses
    sec(layout.PL_OPEX_START_ROW-1, "III.  OPERATING EXPENSES")
    opex = [
        (layout.PL_RM_ROW,       "  R&M Expenses",           "Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_RM_AMOUNT)}"),
        (layout.PL_INS_ROW,      "  Insurance",               "Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_INS_AMOUNT)}"),
        (layout.PL_MARKETING_ROW,"  Marketing Expenses",      "Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_MARKETING_AMOUNT)}"),
        (layout.PL_POWER_ROW,    "  Power & Fuel",            "Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_POWER_AMOUNT)}"),
        (layout.PL_MANPOWER_ROW, "  Manpower / Salaries",     "ManPower", lambda y,c: f"=ManPower!{yl(y)}{layout.mp_annual_row()+2}"),
        (layout.PL_DEPR_ROW,     "  Depreciation",            "Depreciation", lambda y,c: f"=Depreciation!{get_column_letter(layout.DEP_COL_YEAR1+y-1)}{layout.dep_total_depr_row()}"),
        (layout.PL_SGA_ROW,      "  Selling, General & Admin","Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_SGA_AMOUNT)}"),
        (layout.PL_TRANSPORT_ROW,"  Transportation",          "Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_TRANSPORT_AMOUNT)}"),
        (layout.PL_MISC_ROW,     "  Miscellaneous",           "Expenses", lambda y,c: f"=Expenses!{yl(y)}{layout.exp_oh_row(EXP.EXP_OH_MISC_AMOUNT)}"),
    ]
    for row, label, sch, fn in opex:
        lbl(row, label, sch=sch); fr(row, fn)

    ws.row_dimensions[layout.PL_TOTAL_OPEX_ROW].height = 17
    lbl(layout.PL_TOTAL_OPEX_ROW, "Total Operating Expenses", bold=True)
    fr(layout.PL_TOTAL_OPEX_ROW, lambda y,c: f"=SUM({c}{layout.PL_RM_ROW}:{c}{layout.PL_MISC_ROW})", bold=True, bg="F2F2F2")

    # EBIT
    ws.row_dimensions[layout.PL_EBIT_ROW].height = 17
    lbl(layout.PL_EBIT_ROW, "EARNINGS BEFORE INTEREST & TAX  (EBIT)", bold=True, color="1F3864")
    fr(layout.PL_EBIT_ROW, lambda y,c: f"={c}{layout.PL_GROSS_PROFIT_ROW}-{c}{layout.PL_TOTAL_OPEX_ROW}", bold=True, bg="EBF5FB")

    # IV. Finance Costs
    sec(layout.PL_TL_INTEREST_ROW-1, "IV.  FINANCE COSTS")
    lbl(layout.PL_TL_INTEREST_ROW, "  Term Loan Interest", sch="Term Loan")
    fr(layout.PL_TL_INTEREST_ROW, lambda y,c: f"='Term Loan'!{get_column_letter(EXP.TL_ANNUAL_COL_Y1+y-1)}{EXP.TL_ANNUAL_INTEREST_ROW}")
    lbl(layout.PL_WC_INTEREST_ROW, "  Working Capital Interest", sch="W Cap")
    fr(layout.PL_WC_INTEREST_ROW, lambda y,c: f"='W Cap'!{get_column_letter(layout.wcap_col(y))}{layout.WCAP_WC_INTEREST}")
    lbl(layout.PL_TOTAL_FINANCE_ROW, "Total Finance Costs", bold=True)
    fr(layout.PL_TOTAL_FINANCE_ROW, lambda y,c: f"={c}{layout.PL_TL_INTEREST_ROW}+{c}{layout.PL_WC_INTEREST_ROW}", bold=True)

    # PBT
    ws.row_dimensions[layout.PL_PBT_ROW].height = 17
    lbl(layout.PL_PBT_ROW, "PROFIT BEFORE TAX  (PBT)", bold=True, color="1F3864")
    fr(layout.PL_PBT_ROW, lambda y,c: f"={c}{layout.PL_EBIT_ROW}-{c}{layout.PL_TOTAL_FINANCE_ROW}", bold=True, bg="EBF5FB")

    # V. Tax
    sec(layout.PL_TAX_ROW-1, "V.  TAX")
    lbl(layout.PL_TAX_ROW, "  Less: Current Tax", sch="Tax")
    fr(layout.PL_TAX_ROW, lambda y,c: f"=Tax!{yl(y)}11")

    # PAT
    ws.row_dimensions[layout.PL_PAT_ROW].height = 17
    lbl(layout.PL_PAT_ROW, "PROFIT AFTER TAX  (PAT)", bold=True, color="1F3864")
    fr(layout.PL_PAT_ROW, lambda y,c: f"={c}{layout.PL_PBT_ROW}-{c}{layout.PL_TAX_ROW}", bold=True, bg="EBF5FB")

    # Drawings
    drawings_ref = layout.asmp_ref("opex.drawings_base")
    drawings_esc = layout.asmp_ref("opex.drawings_escalation")
    lbl(layout.PL_DRAWINGS_ROW, "  Less: Drawings / Proprietor Withdrawals")
    fr(layout.PL_DRAWINGS_ROW,
       lambda y,c: f"={drawings_ref}*(1+{drawings_esc})^{y-1}" if y > 1 else f"={drawings_ref}")

    # Retained
    lbl(layout.PL_RETAINED_ROW, "  Net Retained Profit (PAT less Drawings)")
    fr(layout.PL_RETAINED_ROW, lambda y,c: f"={c}{layout.PL_PAT_ROW}-{c}{layout.PL_DRAWINGS_ROW}")

    # EBITDA
    ws.row_dimensions[layout.PL_EBITDA_ROW].height = 17
    lbl(layout.PL_EBITDA_ROW, "EBITDA  (EBIT + Depreciation)", bold=True)
    fr(layout.PL_EBITDA_ROW, lambda y,c: f"={c}{layout.PL_EBIT_ROW}+{c}{layout.PL_DEPR_ROW}", bold=True, bg="F2F2F2")

    # EBITDA Margin
    lbl(layout.PL_EBITDA_MARGIN_ROW, "  EBITDA Margin  (EBITDA / Revenue)")
    fr(layout.PL_EBITDA_MARGIN_ROW, lambda y,c: f"=IFERROR({c}{layout.PL_EBITDA_ROW}/{c}{layout.PL_REVENUE_ROW},0)", fmt=FMT_PCT)

    # ICR — Interest Coverage Ratio
    lbl(layout.PL_ICR_ROW, "  Interest Coverage Ratio  (EBIT / Finance Costs)", bold=True)
    fr(layout.PL_ICR_ROW, lambda y,c: f"=IFERROR({c}{layout.PL_EBIT_ROW}/{c}{layout.PL_TOTAL_FINANCE_ROW},0)", fmt="0.00", bold=True, bg="F2F2F2")

    freeze_header(ws, row=5, col=layout.PL_COL_YEAR1)
    return ws
