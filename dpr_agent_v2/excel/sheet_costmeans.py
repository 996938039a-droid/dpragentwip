"""excel/sheet_costmeans.py - Project Cost and Means of Finance.
Category summary uses DIRECT ROW REFERENCES (no SUMIF).
"""
from __future__ import annotations
from openpyxl import Workbook
from core.session_store import SessionStore, AssetCategory
from core.layout_engine import LayoutEngine
from excel.styles import write_title, set_col_widths, FONT_BODY, FONT_FORMULA, FMT_LAKHS, font, fill

ASSET_DATA_START   = 6
FINANCE_HEADER_GAP = 3
CAT_SUMMARY_START  = 80

CAT_ROW = {
    "plant_machinery": 0,
    "civil_works":     1,
    "furniture":       2,
    "vehicle":         3,
    "electrical":      4,
    "other":           5,
}

ASSETCAT_TO_KEY = {
    AssetCategory.PLANT_MACHINERY: "plant_machinery",
    AssetCategory.CIVIL_WORKS:     "civil_works",
    AssetCategory.FURNITURE:       "furniture",
    AssetCategory.VEHICLE:         "vehicle",
    AssetCategory.ELECTRICAL:      "electrical",
    AssetCategory.PRE_OPERATIVE:   "other",
    AssetCategory.OTHER:           "other",
}


def cat_summary_row(cat_key):
    return CAT_SUMMARY_START + CAT_ROW.get(cat_key, 5)


def write_costmeans_sheet(wb, store, layout):
    ws = wb.create_sheet("Cost & Means", 1)
    set_col_widths(ws, {"A": 5, "B": 44, "C": 24, "D": 16, "E": 14})
    write_title(ws, store.project_profile.company_name, "Project Cost & Means of Finance")
    ws.cell(row=3, column=2, value="(All amounts in INR Lakhs)").font = font(size=9, color="808080")

    # Section A: Project Cost
    _hdr(ws, 4, "A  |  PROJECT COST")
    for col, lbl in [(2, "Asset / Item"), (3, "Category"), (4, "Amount (INR Lakhs)")]:
        ws.cell(row=5, column=col, value=lbl).font = font(bold=True, size=9, color="808080")

    assets = store.capital_means.assets
    for i, asset in enumerate(assets):
        r = ASSET_DATA_START + i
        ws.cell(row=r, column=2, value=asset.name).font = FONT_BODY
        ws.cell(row=r, column=3, value=asset.category.value).font = font(size=9, color="606060")
        c = ws.cell(row=r, column=4, value=asset.cost_lakhs)
        c.font = FONT_BODY
        c.number_format = FMT_LAKHS

    first_r = ASSET_DATA_START
    last_r  = ASSET_DATA_START + len(assets) - 1
    total_cost_row = last_r + 2
    _tot(ws, total_cost_row, "TOTAL PROJECT COST", "=SUM(D" + str(first_r) + ":D" + str(last_r) + ")")

    # Section B: Means of Finance
    fin_header_row = total_cost_row + FINANCE_HEADER_GAP
    _hdr(ws, fin_header_row, "B  |  MEANS OF FINANCE")
    fin_data_start = fin_header_row + 1
    sources = store.capital_means.finance_sources

    for i, src in enumerate(sources):
        r = fin_data_start + i
        if src.is_term_loan:
            detail = "@ " + str(round(src.interest_rate*100,1)) + "% p.a., " + str(src.tenor_months) + "mo, " + str(src.moratorium_months) + "mo mora"
        elif src.is_od:
            detail = "OD/CC @ " + str(round(src.interest_rate*100,1)) + "% p.a.  [Working Capital -- not project cost]"
        elif src.is_equity:
            detail = "Promoter contribution"
        else:
            detail = ""
        ws.cell(row=r, column=2, value=src.label).font = FONT_BODY
        ws.cell(row=r, column=3, value=detail).font = font(size=9, color="606060")
        c = ws.cell(row=r, column=4, value=src.amount_lakhs)
        c.font = FONT_BODY
        c.number_format = FMT_LAKHS

    # Project finance = TL + Equity only; OD is working capital, excluded from balance check
    proj_rows = [fin_data_start + i for i, s in enumerate(sources) if not s.is_od]
    od_rows   = [fin_data_start + i for i, s in enumerate(sources) if s.is_od]
    proj_sum  = "+".join("D" + str(r) for r in proj_rows) if proj_rows else "0"

    total_fin_row = fin_data_start + len(sources) + 1
    _tot(ws, total_fin_row, "TOTAL PROJECT FINANCE  (Term Loan + Equity)", "=" + proj_sum)

    if od_rows:
        orow = total_fin_row + 1
        ws.cell(row=orow, column=2, value="  Working Capital / OD  (revolving -- not project cost)").font = font(size=9, italic=True, color="808080")
        c = ws.cell(row=orow, column=4, value="=D" + str(od_rows[0]))
        c.font = font(size=9, italic=True, color="808080")
        c.number_format = FMT_LAKHS
        bal_row = orow + 1
    else:
        bal_row = total_fin_row + 1

    ws.cell(row=bal_row, column=2, value="Balance Check  (Project Finance vs Project Cost)").font = font(size=9, italic=True, color="808080")
    ok  = chr(10003) + " BALANCED"
    bad = chr(10007) + " GAP="
    c = ws.cell(row=bal_row, column=4,
                value='=IF(ROUND(D' + str(total_cost_row) + ',0)=ROUND(D' + str(total_fin_row) + ',0),"' + ok + '","' + bad + '"&TEXT(D' + str(total_fin_row) + '-D' + str(total_cost_row) + ',"#,##0.00"))')
    c.font = font(bold=True, color="006400")

    # Category summary: DIRECT ROW REFERENCES, no SUMIF, no string matching
    cat_asset_rows = {}
    for i, asset in enumerate(assets):
        k = ASSETCAT_TO_KEY.get(asset.category, "other")
        if k not in cat_asset_rows:
            cat_asset_rows[k] = []
        cat_asset_rows[k].append(ASSET_DATA_START + i)

    ws.cell(row=CAT_SUMMARY_START - 2, column=2,
            value="Category Totals  (referenced by Depreciation -- do not move these rows)").font = font(size=9, italic=True, color="808080")

    for cat_key, offset in CAT_ROW.items():
        row = CAT_SUMMARY_START + offset
        rows_for_cat = cat_asset_rows.get(cat_key, [])
        if rows_for_cat:
            formula = "=" + "+".join("D" + str(r) for r in rows_for_cat)
        else:
            formula = "=0"
        ws.cell(row=row, column=2, value=cat_key.replace("_", " ").title()).font = font(size=9, color="808080")
        c = ws.cell(row=row, column=4, value=formula)
        c.font = FONT_FORMULA
        c.number_format = FMT_LAKHS

    return ws


def _hdr(ws, row, text):
    c = ws.cell(row=row, column=2, value=text)
    c.font = font(bold=True, color="FFFFFF")
    c.fill = fill("1F3864")
    ws.row_dimensions[row].height = 16
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)


def _tot(ws, row, label, formula):
    ws.row_dimensions[row].height = 17
    lc = ws.cell(row=row, column=2, value=label)
    lc.font = font(bold=True, color="FFFFFF")
    lc.fill = fill("1F3864")
    vc = ws.cell(row=row, column=4, value=formula)
    vc.font = font(bold=True, color="FFFFFF")
    vc.fill = fill("1F3864")
    vc.number_format = FMT_LAKHS
