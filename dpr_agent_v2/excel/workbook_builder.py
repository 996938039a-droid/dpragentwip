"""
excel/workbook_builder.py
═══════════════════════════
Orchestrates all sheet writers in the correct order.
Returns a completed openpyxl Workbook ready to save/download.
"""

import io
from openpyxl import Workbook

from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.assumption_writer   import write_assumption_sheet
from core.layout_engine         import LayoutEngine
from core.industry_config      import get_profile, applicable_wc_items, applicable_bs_items
from excel.sheet_costmeans     import write_costmeans_sheet
from excel.sheet_revenue       import write_revenue_sheet
from excel.sheet_expenses      import write_expenses_sheet
from excel.sheet_manpower      import write_manpower_sheet
from excel.sheet_depreciation  import write_depreciation_sheet
from excel.sheet_termloan      import write_termloan_sheet
from excel.sheet_wcap          import write_wcap_sheet
from excel.sheet_tax           import write_tax_sheet
from excel.sheet_pl            import write_pl_sheet
from excel.sheet_bs            import write_bs_sheet
from excel.sheet_cfs           import write_cfs_sheet
from excel.sheet_ratio         import write_ratio_sheet


def build_workbook(store: SessionStore) -> Workbook:
    """
    Build the complete DPR workbook from a SessionStore.
    Sheet order matches the DPR format expected by bankers.
    """
    wb = Workbook()
    # Remove the default blank sheet
    del wb[wb.sheetnames[0]]

    # Ensure minimum 1 for LayoutEngine even if no materials (e.g. service business)
    layout = LayoutEngine(
        n_products  = max(1, len(store.revenue_model.products)),
        n_materials = max(1, len(store.cost_structure.raw_materials)),
        n_employees = max(1, len(store.manpower.categories)),
        n_years     = store.projection_years,
    )

    # Sheet build order matters — later sheets reference earlier ones
    # Assumption must be first (all other sheets pull from it)
    write_assumption_sheet(wb, store, layout)   # sheet 1
    write_costmeans_sheet(wb, store, layout)      # sheet 2 — referenced by Depr & CFS

    # Supporting schedules (no inter-schedule dependencies)
    write_revenue_sheet(wb, store, layout)       # sheet 2
    write_manpower_sheet(wb, store, layout)      # sheet 3
    write_depreciation_sheet(wb, store, layout)  # sheet 4
    write_termloan_sheet(wb, store, layout)      # sheet 5
    write_wcap_sheet(wb, store, layout)          # sheet 6

    # Expenses references Revenue + Depreciation
    write_expenses_sheet(wb, store, layout)      # sheet 7

    # Tax references PL (circular ref avoided — Tax reads PBT, PL reads Tax)
    # We write Tax before PL but PL writes first; Excel resolves on open
    write_tax_sheet(wb, store, layout)           # sheet 8

    # Financial statements (P&L must come before BS and CFS)
    write_pl_sheet(wb, store, layout)            # sheet 9
    write_cfs_sheet(wb, store, layout)           # sheet 10
    write_bs_sheet(wb, store, layout)            # sheet 11
    write_ratio_sheet(wb, store, layout)         # sheet 12

    # ── Industry intelligence: hide inapplicable rows ───────────────────────
    _apply_industry_visibility(wb, store, layout)

    return wb


def build_workbook_bytes(store: SessionStore) -> bytes:
    """Build the workbook and return as bytes for Streamlit download."""
    wb = build_workbook(store)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def suggested_filename(store: SessionStore) -> str:
    """Generate a clean filename for the DPR."""
    from datetime import date
    company = store.project_profile.company_name
    safe    = "".join(c if c.isalnum() or c in " _-" else "_" for c in company)
    safe    = safe.replace(" ", "_")[:40]
    today   = date.today().strftime("%Y%m%d")
    return f"DPR_{safe}_{today}.xlsx"


def _apply_industry_visibility(wb, store, layout):
    """
    Hide rows that are not applicable based on store.flags (live FieldFlags).
    Hides rows without deleting formulas — balance sheet stays balanced.
    Flags are 1=show, 0=hide. Drawings is ALWAYS 1 (never hidden).
    """
    from openpyxl.utils import get_column_letter

    # Use the live field flags from store (may have been updated by handlers)
    f = store.flags

    # ── W Cap sheet — hide inapplicable stock rows ───────────────────────────
    if "W Cap" in wb.sheetnames:
        ws_wc = wb["W Cap"]
        _hide(ws_wc, layout.WCAP_STOCK_WIP,      f.wip          == 0)
        _hide(ws_wc, layout.WCAP_STOCK_FG,        f.finished_goods == 0)
        _hide(ws_wc, layout.WCAP_STOCK_COLD,      f.cold_store   == 0)
        _hide(ws_wc, layout.WCAP_STOCK_RM,        f.raw_materials == 0)
        _hide(ws_wc, layout.WCAP_CREDITORS_RM,    f.raw_materials == 0)

    # ── BS sheet — hide inapplicable line items ───────────────────────────────
    if "BS" in wb.sheetnames:
        ws_bs = wb["BS"]
        _hide(ws_bs, layout.BS_WIP_ROW,           f.wip          == 0)
        _hide(ws_bs, layout.BS_FG_ROW,            f.finished_goods == 0)
        _hide(ws_bs, layout.BS_COLD_STORE_ROW,    f.cold_store   == 0)
        _hide(ws_bs, layout.BS_INTANGIBLE_ROW,    f.intangibles  == 0)
        _hide(ws_bs, layout.BS_SECURITY_DEP_ROW,  f.security_deposits == 0)
        _hide(ws_bs, layout.BS_NON_CURR_INV_ROW,  f.nc_investments == 0)

    # ── PL sheet ──────────────────────────────────────────────────────────────
    if "PL" in wb.sheetnames:
        ws_pl = wb["PL"]
        # drawings is ALWAYS visible (flag is always 1)
        _hide(ws_pl, layout.PL_COGS_ROW,       f.rm_cogs   == 0)
        _hide(ws_pl, layout.PL_COGS_ROW + 1,   f.rm_cogs   == 0)
        _hide(ws_pl, layout.PL_TRANSPORT_ROW,  f.transport == 0)
        _hide(ws_pl, layout.PL_POWER_ROW,      f.power     == 0)
        _hide(ws_pl, layout.PL_MARKETING_ROW,  f.marketing == 0)

    # ── Expenses sheet — hide COGS if no raw materials ────────────────────────
    if "Expenses" in wb.sheetnames and f.rm_cogs == 0:
        ws_exp = wb["Expenses"]
        for r in range(layout.EXP_COGS_START, layout.exp_total_cogs_row() + 1):
            ws_exp.row_dimensions[r].hidden = True


def _hide(ws, row: int, condition: bool):
    """Hide a row if condition is True."""
    if condition:
        ws.row_dimensions[row].hidden = True
