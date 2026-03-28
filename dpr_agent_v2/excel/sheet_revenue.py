"""
excel/sheet_revenue.py
════════════════════════
Revenue sheet — computes annual production volumes and revenues
for all products across all projection years.

Formula logic per product per year:
  Production = working_days × months × capacity/day × output_ratio × split% × utilisation%
  Price      = base_price × (1 + escalation)^(year-1)
  Revenue    = Production × Price / 100000   (convert to lakhs)

All parameters reference Assumption sheet via layout.asmp_ref().
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header, write_header_row,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_SUBHEADER, FONT_REF,
    FILL_HEADER, FILL_TOTAL, FILL_ALT, FILL_WHITE,
    AL_LEFT, AL_CENTER, AL_RIGHT,
    FMT_LAKHS, FMT_INT, FMT_DATE, FMT_ZERO,
    font, fill, border
)


def write_revenue_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("Revenue")

    set_col_widths(ws, {
        "A": 4, "B": 38, "C": 14,
        **{get_column_letter(layout.year_col(y)): 13 for y in range(1, layout.n_years + 1)}
    })

    write_title(ws, store.project_profile.company_name, "Statement of Total Revenue")

    # ── Row 4: year headers ────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Particulars").font  = font(bold=True)
    ws.cell(row=4, column=3, value="Basis").font        = font(bold=True, size=9, color="808080")

    op_start = store.project_profile.operation_start_date or "2026-04"
    year_str, month_str = op_start.split("-")
    start_year, start_month = int(year_str), int(month_str)

    for y in range(1, layout.n_years + 1):
        col = layout.year_col(y)
        # Calculate end date of this financial year
        end_month = start_month - 1 if start_month > 1 else 12
        end_year  = start_year + y if start_month > 1 else start_year + y - 1
        import datetime
        try:
            end_date = datetime.date(end_year, end_month, 
                                     [31,28,31,30,31,30,31,31,30,31,30,31][end_month-1])
        except Exception:
            end_date = f"Year {y}"
        c = ws.cell(row=4, column=col, value=f"Year {y}" if isinstance(end_date, str) else end_date)
        c.font      = font(bold=True, color="FFFFFF")
        c.fill      = fill("1F3864")
        c.alignment = AL_CENTER
        if not isinstance(end_date, str):
            c.number_format = "MMM-YYYY"

    # ── Rows 5–10: operational parameters ─────────────────────────────────────
    params = [
        (5,  "Months in Operation",      "months",     lambda y: "=12"),
        (6,  "Working Days per Month",   "days",       lambda y: f"={layout.asmp_ref('capacity.working_days')}"),
        (7,  "Total Capacity per Day",   "units/day",  None),   # sum of all products
        (8,  "Output Ratio",             "fraction",   None),   # N/A for bats (=1)
        (10, "Capacity Utilisation %",   "fraction",   None),   # dynamic formula
    ]

    ws.cell(row=5, column=2, value="Months in Operation").font = FONT_BODY
    ws.cell(row=5, column=3, value="months").font = font(size=9, color="808080")
    ws.cell(row=6, column=2, value="Working Days per Month").font = FONT_BODY
    ws.cell(row=6, column=3, value="days").font = font(size=9, color="808080")
    ws.cell(row=10, column=2, value="Capacity Utilisation %").font = FONT_BODY
    ws.cell(row=10, column=3, value="fraction").font = font(size=9, color="808080")

    for y in range(1, layout.n_years + 1):
        col = layout.year_col(y)
        col_l = layout.year_col_letter(y)

        # Months in operation = 12
        ws.cell(row=5, column=col, value="=12").font = FONT_FORMULA

        # Working days
        ws.cell(row=6, column=col,
                value=f"={layout.asmp_ref('capacity.working_days')}").font = FONT_REF

        # Utilisation ramp
        if y == 1:
            ws.cell(row=10, column=col,
                    value=f"={layout.asmp_ref('capacity.year1_util')}").font = FONT_REF
        else:
            prev_col = layout.year_col_letter(y - 1)
            ws.cell(row=10, column=col,
                    value=(f"=IF({prev_col}10<{layout.asmp_ref('capacity.max_util')},"
                           f"{prev_col}10+{layout.asmp_ref('capacity.annual_increment')},"
                           f"{prev_col}10)")).font = FONT_FORMULA

    # ── Product blocks ─────────────────────────────────────────────────────────
    for p_idx, product in enumerate(store.revenue_model.products):
        base_row     = layout.rev_product_start(p_idx)
        prod_row     = layout.rev_production_row(p_idx)   # base+1
        price_row    = layout.rev_price_row(p_idx)         # base+4
        revenue_row  = layout.rev_revenue_row(p_idx)       # base+5

        # Sub-header: product name
        ws.row_dimensions[base_row].height = 16
        ph = ws.cell(row=base_row, column=2, value=f"  {product.name}")
        ph.font = font(bold=True, color="1F3864")
        ph.fill = fill("D6DCE4")
        ws.merge_cells(start_row=base_row, start_column=2,
                       end_row=base_row,   end_column=3)

        # Row labels
        ws.cell(row=prod_row,    column=2, value=f"    Total Production").font = FONT_BODY
        ws.cell(row=prod_row,    column=3, value=product.unit).font = font(size=9, color="808080")
        ws.cell(row=prod_row+1,  column=2, value=f"    Output (after ratio)").font = FONT_BODY
        ws.cell(row=prod_row+1,  column=3, value=product.unit).font = font(size=9, color="808080")
        ws.cell(row=price_row,   column=2, value=f"    Price per {product.unit} (₹)").font = FONT_BODY
        ws.cell(row=price_row,   column=3, value=f"₹/{product.unit}").font = font(size=9, color="808080")
        ws.cell(row=revenue_row, column=2, value=f"    {product.name} — Revenue").font = font(bold=True)
        ws.cell(row=revenue_row, column=3, value="₹ Lakhs").font = font(size=9, color="808080")

        # Assumption refs for this product's parameters — NO hardcoding
        name_row = layout.asmp_row("revenue.product.name", p_idx)
        cap_row  = layout.asmp_row("revenue.product.capacity", p_idx)
        cap_ref        = f"Assumption!$E${cap_row}"    # capacity per day
        output_ref     = f"Assumption!$F${name_row}"   # output ratio (col F of name row)
        split_ref      = f"Assumption!$F${cap_row}"    # split % (col F of cap row)

        for y in range(1, layout.n_years + 1):
            col   = layout.year_col(y)
            col_l = layout.year_col_letter(y)

            # Total production = months × days × capacity × output_ratio × split% × utilisation
            # All parameters from Assumption sheet — zero hardcoding
            ws.cell(row=prod_row, column=col,
                    value=f"={col_l}5*{col_l}6*{cap_ref}*{output_ref}*{split_ref}*{col_l}10").font = FONT_FORMULA
            ws.cell(row=prod_row, column=col).number_format = FMT_INT

            # Output after ratio (same for bats — ratio=1)
            ws.cell(row=prod_row+1, column=col,
                    value=f"={col_l}{prod_row}").font = FONT_FORMULA
            ws.cell(row=prod_row+1, column=col).number_format = FMT_INT

            # Price with escalation
            price_ref = layout.asmp_ref("revenue.product.price", p_idx)
            esc_ref   = layout.asmp_ref("revenue.product.escalation", p_idx)
            if y == 1:
                ws.cell(row=price_row, column=col,
                        value=f"={price_ref}").font = FONT_REF
            else:
                prev_col = layout.year_col_letter(y - 1)
                ws.cell(row=price_row, column=col,
                        value=f"={prev_col}{price_row}*(1+{esc_ref})").font = FONT_FORMULA
            ws.cell(row=price_row, column=col).number_format = FMT_INT

            # Revenue = production × price / 100000
            ws.cell(row=revenue_row, column=col,
                    value=f"=({col_l}{prod_row+1}*{col_l}{price_row})/100000").font = FONT_FORMULA
            ws.cell(row=revenue_row, column=col).number_format = FMT_LAKHS
            ws.cell(row=revenue_row, column=col).fill = fill("EBF5FB")

    # ── Total Revenue row ──────────────────────────────────────────────────────
    total_row = layout.rev_total_row()
    ws.row_dimensions[total_row].height = 18
    t = ws.cell(row=total_row, column=2, value="TOTAL REVENUE FROM OPERATIONS")
    t.font = FONT_TOTAL
    t.fill = FILL_TOTAL
    ws.cell(row=total_row, column=3, value="₹ Lakhs").font = font(bold=True, size=9)

    for y in range(1, layout.n_years + 1):
        col = layout.year_col(y)
        rev_rows = [layout.rev_revenue_row(i) for i in range(layout.n_products)]
        sum_formula = "+".join(
            f"{layout.year_col_letter(y)}{r}" for r in rev_rows
        )
        c = ws.cell(row=total_row, column=col, value=f"={sum_formula}")
        c.font           = FONT_TOTAL
        c.fill           = FILL_TOTAL
        c.number_format  = FMT_LAKHS

    freeze_header(ws, row=5, col=4)
    return ws
