"""excel/sheet_ratio.py — Ratios: DSCR, ROCE, Break-even, Profitability, IRR, Asset Turnover."""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import write_title, set_col_widths, FONT_BODY, FONT_FORMULA, FMT_LAKHS, FMT_PCT, font, fill

def write_ratio_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("Ratio")
    set_col_widths(ws, {"A": 4, "B": 44,
                        **{get_column_letter(layout.ratio_col(y)): 13 for y in range(1, layout.n_years+1)}})
    write_title(ws, store.project_profile.company_name, "Key Financial Ratios & Indicators")
    ws.cell(row=3, column=2, value="For Banker Appraisal").font = font(size=9, italic=True, color="808080")

    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=2, value="Ratio / Indicator").font = font(bold=True)
    for y in range(1, layout.n_years+1):
        col = layout.ratio_col(y)
        c = ws.cell(row=4, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")

    def lbl(row, text, bold=False):
        ws.cell(row=row, column=2, value=text).font = font(bold=bold)

    def sec(row, text):
        c = ws.cell(row=row, column=2, value=text)
        c.font = font(bold=True, color="FFFFFF"); c.fill = fill("1F3864")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1+layout.n_years+2)

    def fr(row, fn, fmt=FMT_LAKHS, bold=False, bg=None):
        for y in range(1, layout.n_years+1):
            col = layout.ratio_col(y); col_l = get_column_letter(col)
            c = ws.cell(row=row, column=col, value=fn(y, col_l))
            c.font = font(bold=bold); c.number_format = fmt
            if bg: c.fill = fill(bg)

    pl  = layout.pl_col_letter
    bsc = layout.bs_col_letter
    EXP = LayoutEngine
    tl_col = lambda y: get_column_letter(EXP.TL_ANNUAL_COL_Y1+y-1)

    # ── DSCR ─────────────────────────────────────────────────────────────────
    sec(layout.RATIO_DSCR_NUM_ROW-2, "DEBT SERVICE COVERAGE RATIO  (DSCR)")
    lbl(layout.RATIO_DSCR_NUM_ROW, "  Numerator  (PAT + Depreciation + Interest)")
    lbl(layout.RATIO_DSCR_DEN_ROW, "  Denominator  (Principal + Interest on TL)")
    fr(layout.RATIO_DSCR_NUM_ROW,
       lambda y,c: f"=PL!{pl(y)}{layout.PL_PAT_ROW}+PL!{pl(y)}{layout.PL_DEPR_ROW}+PL!{pl(y)}{layout.PL_TOTAL_FINANCE_ROW}")
    fr(layout.RATIO_DSCR_DEN_ROW,
       lambda y,c: f"='Term Loan'!{tl_col(y)}{EXP.TL_ANNUAL_PRINCIPAL_ROW}+'Term Loan'!{tl_col(y)}{EXP.TL_ANNUAL_INTEREST_ROW}")
    lbl(layout.RATIO_DSCR_ROW, "DSCR", bold=True)
    fr(layout.RATIO_DSCR_ROW,
       lambda y,c: f"=IFERROR({c}{layout.RATIO_DSCR_NUM_ROW}/{c}{layout.RATIO_DSCR_DEN_ROW},0)",
       fmt="0.00", bold=True, bg="EBF5FB")
    lbl(layout.RATIO_AVG_DSCR_ROW, "  Average DSCR (all years)")
    all_dscr = ",".join(f"{get_column_letter(layout.ratio_col(y))}{layout.RATIO_DSCR_ROW}" for y in range(1, layout.n_years+1))
    ws.cell(row=layout.RATIO_AVG_DSCR_ROW, column=layout.ratio_col(1), value=f"=AVERAGE({all_dscr})").number_format = "0.00"

    # ── ROCE ─────────────────────────────────────────────────────────────────
    sec(layout.RATIO_ROCE_ROW-2, "RETURN ON CAPITAL EMPLOYED  (ROCE)")
    lbl(layout.RATIO_ROCE_ROW, "ROCE  (EBIT / Total Assets)", bold=True)
    fr(layout.RATIO_ROCE_ROW,
       lambda y,c: f"=IFERROR(PL!{pl(y)}{layout.PL_EBIT_ROW}/BS!{bsc(y)}{layout.BS_TOTAL_ASSETS_ROW},0)",
       fmt=FMT_PCT, bold=True, bg="EBF5FB")
    lbl(layout.RATIO_ROCE_ROW+1, "  Average ROCE")
    avg_roce = ",".join(f"{get_column_letter(layout.ratio_col(y))}{layout.RATIO_ROCE_ROW}" for y in range(1, layout.n_years+1))
    ws.cell(row=layout.RATIO_ROCE_ROW+1, column=layout.ratio_col(1), value=f"=AVERAGE({avg_roce})").number_format = FMT_PCT

    # ── Break-even ───────────────────────────────────────────────────────────
    sec(layout.RATIO_BEP_FIXED_ROW-2, "BREAK-EVEN POINT ANALYSIS")
    lbl(layout.RATIO_BEP_FIXED_ROW, "  Fixed Costs (Opex + Interest + Depreciation)")
    lbl(layout.RATIO_BEP_CONTRIB,   "  Contribution  (Revenue – COGS)")
    lbl(layout.RATIO_BEP_ROW,       "BREAK-EVEN  (% of Revenue)", bold=True)
    fr(layout.RATIO_BEP_FIXED_ROW,
       lambda y,c: (f"=PL!{pl(y)}{layout.PL_TOTAL_OPEX_ROW}-PL!{pl(y)}{layout.PL_COGS_ROW}"
                    f"+PL!{pl(y)}{layout.PL_TOTAL_FINANCE_ROW}+PL!{pl(y)}{layout.PL_DEPR_ROW}"))
    fr(layout.RATIO_BEP_CONTRIB,
       lambda y,c: f"=PL!{pl(y)}{layout.PL_REVENUE_ROW}-PL!{pl(y)}{layout.PL_COGS_ROW}")
    fr(layout.RATIO_BEP_ROW,
       lambda y,c: f"=IFERROR({c}{layout.RATIO_BEP_FIXED_ROW}/{c}{layout.RATIO_BEP_CONTRIB},0)",
       fmt=FMT_PCT, bold=True, bg="EBF5FB")

    # ── Profitability ─────────────────────────────────────────────────────────
    sec(layout.RATIO_OP_MARGIN-2, "PROFITABILITY RATIOS")
    lbl(layout.RATIO_OP_MARGIN,   "  Operating Margin  (EBIT / Revenue)")
    lbl(layout.RATIO_NET_MARGIN,  "  Net Profit Margin  (PAT / Revenue)")
    lbl(layout.RATIO_DEBT_EQUITY, "  Debt / Equity  (TL / Total Equity)")
    lbl(layout.RATIO_ASSET_TURNOVER, "  Asset Turnover  (Revenue / Total Assets)", bold=True)
    fr(layout.RATIO_OP_MARGIN,
       lambda y,c: f"=IFERROR(PL!{pl(y)}{layout.PL_EBIT_ROW}/PL!{pl(y)}{layout.PL_REVENUE_ROW},0)", fmt=FMT_PCT)
    fr(layout.RATIO_NET_MARGIN,
       lambda y,c: f"=IFERROR(PL!{pl(y)}{layout.PL_PAT_ROW}/PL!{pl(y)}{layout.PL_REVENUE_ROW},0)", fmt=FMT_PCT)
    fr(layout.RATIO_DEBT_EQUITY,
       lambda y,c: f"=IFERROR('Term Loan'!{tl_col(y)}{EXP.TL_ANNUAL_CLOSING_ROW}/BS!{bsc(y)}{layout.BS_TOTAL_EQUITY_ROW},0)", fmt="0.00")
    fr(layout.RATIO_ASSET_TURNOVER,
       lambda y,c: f"=IFERROR(PL!{pl(y)}{layout.PL_REVENUE_ROW}/BS!{bsc(y)}{layout.BS_TOTAL_ASSETS_ROW},0)",
       fmt="0.00", bold=True, bg="EBF5FB")

    # ── Project IRR ───────────────────────────────────────────────────────────
    sec(layout.RATIO_IRR_CF_ROW-2, "PROJECT IRR  (Free Cash Flow Basis)")
    lbl(layout.RATIO_IRR_CF_ROW,  "  Free Cash Flows  (PAT + Depreciation)")
    lbl(layout.RATIO_IRR_ROW,     "PROJECT IRR", bold=True)

    # Year 0 cash flow in the label column
    from excel.sheet_costmeans import ASSET_DATA_START
    n_assets = len(store.capital_means.assets)
    cm_total_row = ASSET_DATA_START + n_assets + 1
    # Place Year 0 (initial investment) one column before Year 1
    y0_col = layout.ratio_col(1) - 1  # column before Year 1
    ws.cell(row=layout.RATIO_IRR_CF_ROW-1, column=y0_col, value="Year 0").font = font(bold=True, size=9, color="808080")
    ws.cell(row=layout.RATIO_IRR_CF_ROW,   column=y0_col,
            value=f"=-'Cost & Means'!$D${cm_total_row}").number_format = FMT_LAKHS

    # Free cash flows Years 1-7
    fr(layout.RATIO_IRR_CF_ROW,
       lambda y,c: f"=PL!{pl(y)}{layout.PL_PAT_ROW}+PL!{pl(y)}{layout.PL_DEPR_ROW}")

    # IRR formula: range from Y0 col to last year col
    y0_col_l   = get_column_letter(y0_col)
    last_col_l = get_column_letter(layout.ratio_col(layout.n_years))
    irr_range  = f"{y0_col_l}{layout.RATIO_IRR_CF_ROW}:{last_col_l}{layout.RATIO_IRR_CF_ROW}"
    # IRR: IFERROR returns "N/A" string when all cash flows are negative
    c = ws.cell(row=layout.RATIO_IRR_ROW, column=layout.ratio_col(1),
                value=f'=IFERROR(TEXT(IRR({irr_range}),"0.00%"),"N/A")')
    c.font = font(bold=True); c.number_format = "@"; c.fill = fill("EBF5FB")

    return ws
