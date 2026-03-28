"""
excel/sheet_expenses.py
═════════════════════════
Computes all operating expenses:
  A. Cost of Sales — raw material costs
  B. Operating Overhead Expenses

KEY FIXES vs v1:
  1. NO ×1000 multiplier — production is already in output units (pieces/kg/L),
     not in tons. Quantity = Total Production × input_per_output_unit
  2. ALL Assumption cell references use layout.asmp_ref() — never hardcoded rows
  3. Material price and escalation reference their own correct rows per material
  4. Overhead rows reference fixed opex rows, not material rows
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths, freeze_header,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_REF,
    FILL_TOTAL, AL_LEFT, AL_CENTER,
    FMT_LAKHS, FMT_INT, FMT_ZERO,
    font, fill
)


def write_expenses_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("Expenses")

    set_col_widths(ws, {
        "A": 4, "B": 40, "C": 14,
        **{get_column_letter(layout.year_col(y)): 13
           for y in range(1, layout.n_years + 1)}
    })

    write_title(ws, store.project_profile.company_name,
                "Statement of Calculation of Expenses")

    # Year headers row 3
    ws.row_dimensions[3].height = 18
    ws.cell(row=3, column=2, value="Particulars").font = font(bold=True)
    ws.cell(row=3, column=3, value="Basis").font       = font(bold=True, size=9, color="808080")
    for y in range(1, layout.n_years + 1):
        col = layout.year_col(y)
        c = ws.cell(row=3, column=col, value=f"Year {y}")
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    # ── Section A: Cost of Sales — Raw Materials ───────────────────────────────
    sec_a_row = 4
    ws.row_dimensions[sec_a_row].height = 16
    s = ws.cell(row=sec_a_row, column=2, value="A  |  COST OF SALES — RAW MATERIALS")
    s.font = font(bold=True, color="FFFFFF")
    s.fill = fill("2E75B6")
    ws.merge_cells(start_row=sec_a_row, start_column=2,
                   end_row=sec_a_row,   end_column=3 + layout.n_years)

    # One block per raw material: qty row, price row, cost row
    for m_idx, mat in enumerate(store.cost_structure.raw_materials):
        qty_row   = layout.exp_qty_row(m_idx)
        price_row = layout.exp_price_row(m_idx)
        cost_row  = layout.exp_cost_row(m_idx)

        # Labels
        ws.cell(row=qty_row,   column=2,
                value=f"  {mat.name} — Quantity Used").font = FONT_BODY
        ws.cell(row=qty_row,   column=3, value=mat.unit).font = font(size=9, color="808080")
        ws.cell(row=price_row, column=2,
                value=f"  {mat.name} — Price per {mat.unit}").font = FONT_BODY
        ws.cell(row=price_row, column=3,
                value=f"₹/{mat.unit}").font = font(size=9, color="808080")
        ws.cell(row=cost_row,  column=2,
                value=f"  {mat.name} — Cost").font = font(bold=True)
        ws.cell(row=cost_row,  column=3, value="₹ Lakhs").font = font(size=9, color="808080")
        ws.cell(row=cost_row,  column=2).fill = fill("EBF5FB")

        # Get the input_per_output for this material
        input_per_output = mat.input_per_output

        # Assumption refs for this material
        price_ref = layout.asmp_ref("material.price", m_idx)
        esc_ref   = layout.asmp_ref("material.escalation", m_idx)

        # Determine which products this material applies to
        # applies_to=None means shared (all products); list means product-specific
        applies_to = mat.applies_to
        if applies_to is None:
            applicable = list(range(layout.n_products))
        else:
            applicable = [
                p for p, prod in enumerate(store.revenue_model.products)
                if prod.name in applies_to
            ]
            if not applicable:
                applicable = list(range(layout.n_products))  # safe fallback

        for y in range(1, layout.n_years + 1):
            col     = layout.year_col(y)
            col_l   = layout.year_col_letter(y)
            rev_col = layout.year_col_letter(y)

            # Quantity = sum of APPLICABLE product productions × input_per_output
            # input_per_output referenced from Assumption col F — NOT hardcoded
            ipp_ref = f"Assumption!$F${layout.asmp_row('material.price', m_idx)}"
            prod_terms = "+".join(
                f"Revenue!{rev_col}{layout.rev_production_row(p)}"
                for p in applicable
            )
            ws.cell(row=qty_row, column=col,
                    value=f"=({prod_terms})*{ipp_ref}").font = FONT_FORMULA
            ws.cell(row=qty_row, column=col).number_format = FMT_INT


            # Price with escalation from Assumption
            if y == 1:
                ws.cell(row=price_row, column=col,
                        value=f"={price_ref}").font = FONT_REF
            else:
                prev_col = layout.year_col_letter(y - 1)
                ws.cell(row=price_row, column=col,
                        value=f"={prev_col}{price_row}*(1+{esc_ref})").font = FONT_FORMULA
            ws.cell(row=price_row, column=col).number_format = FMT_INT

            # Cost = qty × price / 100000
            ws.cell(row=cost_row, column=col,
                    value=f"=({col_l}{qty_row}*{col_l}{price_row})/100000").font = FONT_FORMULA
            ws.cell(row=cost_row, column=col).number_format = FMT_LAKHS

    # Total COGS row
    total_cogs_row = layout.exp_total_cogs_row()
    ws.row_dimensions[total_cogs_row].height = 17
    t = ws.cell(row=total_cogs_row, column=2, value="TOTAL COST OF SALES")
    t.font = font(bold=True, color="FFFFFF")
    t.fill = fill("2E75B6")
    ws.cell(row=total_cogs_row, column=3, value="₹ Lakhs").font = font(bold=True, size=9)

    for y in range(1, layout.n_years + 1):
        col   = layout.year_col(y)
        col_l = layout.year_col_letter(y)
        cost_rows = [layout.exp_cost_row(m) for m in range(layout.n_materials)]
        total_formula = "+".join(f"{col_l}{r}" for r in cost_rows)
        c = ws.cell(row=total_cogs_row, column=col, value=f"={total_formula}")
        c.font          = font(bold=True, color="FFFFFF")
        c.fill          = fill("2E75B6")
        c.number_format = FMT_LAKHS

    # ── Section B: Operating Overhead Expenses ────────────────────────────────
    # No intermediate reference rows — all overhead formulas reference
    # Revenue and Depreciation sheets directly to avoid any collision risk.
    oh_start = layout.exp_overhead_start()
    sec_b_row = total_cogs_row + 1
    ws.row_dimensions[sec_b_row].height = 16
    s2 = ws.cell(row=sec_b_row, column=2, value="B  |  OPERATING OVERHEAD EXPENSES")
    s2.font = font(bold=True, color="FFFFFF")
    s2.fill = fill("2E75B6")
    ws.merge_cells(start_row=sec_b_row, start_column=2,
                   end_row=sec_b_row,   end_column=3 + layout.n_years)

    # Each overhead item: rate row + amount row
    overheads = [
        # (rate_offset, amt_offset, label, basis_label, rate_key, is_pct_revenue, is_pct_nfa, is_base)
        (LayoutEngine.EXP_OH_RM_RATE,          LayoutEngine.EXP_OH_RM_AMOUNT,
         "Repair & Maintenance", "% of Net Fixed Assets",
         "opex.rm_pct_fa", "opex.rm_escalation", "nfa"),

        (LayoutEngine.EXP_OH_INS_RATE,         LayoutEngine.EXP_OH_INS_AMOUNT,
         "Insurance", "% of Net Fixed Assets",
         "opex.insurance_pct_fa", "opex.insurance_escalation", "nfa"),

        (LayoutEngine.EXP_OH_POWER_RATE,       LayoutEngine.EXP_OH_POWER_AMOUNT,
         "Power & Fuel", "% of Revenue",
         "opex.power_pct_revenue", "opex.power_escalation", "rev"),

        (LayoutEngine.EXP_OH_MARKETING_RATE,   LayoutEngine.EXP_OH_MARKETING_AMOUNT,
         "Marketing Expenses", "% of Revenue",
         "opex.marketing_pct_revenue", "opex.marketing_escalation", "rev"),

        (LayoutEngine.EXP_OH_SGA_BASE,         LayoutEngine.EXP_OH_SGA_AMOUNT,
         "Selling, General & Admin", "₹ Lakhs (escalating)",
         "opex.sga_base", "opex.sga_escalation", "base"),

        (LayoutEngine.EXP_OH_TRANSPORT_BASE,   LayoutEngine.EXP_OH_TRANSPORT_AMOUNT,
         "Transportation", "₹ Lakhs (escalating)",
         "opex.transport_base", "opex.transport_escalation", "base"),

        (LayoutEngine.EXP_OH_MISC_BASE,        LayoutEngine.EXP_OH_MISC_AMOUNT,
         "Miscellaneous", "₹ Lakhs (escalating)",
         "opex.misc_base", "opex.misc_escalation", "base"),
    ]

    for rate_offset, amt_offset, label, basis_lbl, rate_key, esc_key, basis_type in overheads:
        rate_row = layout.exp_oh_row(rate_offset)
        amt_row  = layout.exp_oh_row(amt_offset)

        ws.cell(row=rate_row, column=2,
                value=f"  {label} — Rate").font = font(size=9, color="808080", italic=True)
        ws.cell(row=rate_row, column=3, value=basis_lbl).font = font(size=9, color="808080")
        ws.cell(row=amt_row,  column=2, value=f"  {label}").font = FONT_BODY
        ws.cell(row=amt_row,  column=3, value="₹ Lakhs").font = font(size=9, color="808080")

        rate_ref = layout.asmp_ref(rate_key)
        esc_ref  = layout.asmp_ref(esc_key)

        # Determine correct format for rate row
        rate_fmt = FMT_LAKHS if basis_type == "base" else "0.00%"

        for y in range(1, layout.n_years + 1):
            col   = layout.year_col(y)
            col_l = layout.year_col_letter(y)

            # Rate row — escalating from assumption
            if y == 1:
                ws.cell(row=rate_row, column=col,
                        value=f"={rate_ref}").font = FONT_REF
            else:
                prev_col = layout.year_col_letter(y - 1)
                ws.cell(row=rate_row, column=col,
                        value=f"={prev_col}{rate_row}*(1+{esc_ref})").font = FONT_FORMULA
            ws.cell(row=rate_row, column=col).number_format = rate_fmt

            # Amount row — direct cross-sheet refs (no intermediate ref rows)
            dep_col = get_column_letter(layout.DEP_COL_YEAR1 + y - 1)
            if basis_type == "rev":
                # Direct Revenue reference — no intermediate row needed
                ws.cell(row=amt_row, column=col,
                        value=f"={col_l}{rate_row}*Revenue!{col_l}{layout.rev_total_row()}").font = FONT_FORMULA
            elif basis_type == "nfa":
                # Direct Depreciation NFA reference — no intermediate row needed
                ws.cell(row=amt_row, column=col,
                        value=f"={col_l}{rate_row}*Depreciation!{dep_col}{layout.dep_net_block_row()}").font = FONT_FORMULA
            else:  # base — rate_row already holds the absolute amount
                ws.cell(row=amt_row, column=col,
                        value=f"={col_l}{rate_row}").font = FONT_FORMULA
            ws.cell(row=amt_row, column=col).number_format = FMT_LAKHS

    freeze_header(ws, row=4, col=4)
    return ws
