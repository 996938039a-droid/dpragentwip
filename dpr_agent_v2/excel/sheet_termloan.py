"""
excel/sheet_termloan.py
═════════════════════════
Term Loan monthly amortisation schedule + annual summary.
Flat principal repayment (equal instalments, declining interest).
"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from core.session_store import SessionStore
from core.layout_engine import LayoutEngine
from excel.styles import (
    write_title, set_col_widths,
    FONT_BODY, FONT_FORMULA, FONT_TOTAL, FONT_REF,
    FILL_TOTAL, AL_LEFT, AL_CENTER,
    FMT_LAKHS, FMT_INT,
    font, fill
)


def write_termloan_sheet(wb: Workbook, store: SessionStore, layout: LayoutEngine):
    ws = wb.create_sheet("Term Loan")

    set_col_widths(ws, {"A": 6, "B": 5, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})
    write_title(ws, store.project_profile.company_name,
                "Term Loan Amortisation Schedule")

    tls = store.capital_means.term_loans
    if not tls:
        ws.cell(row=5, column=2, value="No term loan in this project.")
        return ws

    tl = tls[0]
    principal     = tl.amount_lakhs
    rate_pa       = tl.interest_rate
    moratorium    = tl.moratorium_months
    tenor         = tl.tenor_months
    repayment_mth = tenor - moratorium
    emi_principal = principal / repayment_mth if repayment_mth > 0 else 0

    # Assumption references — all parameters from Assumption sheet, not hardcoded
    tl_rate_ref = layout.asmp_ref("finance.tl_rate")      # annual rate
    tl_amt_ref  = layout.asmp_ref("finance.tl_amount")    # loan amount
    tl_rep_ref  = layout.asmp_ref("finance.tl_repayment") # repayment months (derived formula in Assumption)
    monthly_rate_formula = f"{tl_rate_ref}/12"            # e.g. Assumption!$E$171/12

    # ── Monthly schedule ──────────────────────────────────────────────────────
    SCHED_START = 5
    ws.row_dimensions[SCHED_START - 1].height = 16
    for col, label in [(1,"Month"),(2,"Year"),(3,"Opening"),(4,"Principal"),
                       (5,"Interest"),(6,"Total EMI"),(7,"Closing")]:
        c = ws.cell(row=SCHED_START - 1, column=col, value=label)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    current_year = 1
    months_in_year = 0

    for month in range(1, tenor + 1):
        row = SCHED_START + month - 1
        ws.row_dimensions[row].height = 14

        # Year label = calendar year (1-based, 12 months per year)
        # Months 1-12 = Year 1, Months 13-24 = Year 2, etc.
        # Moratorium months and repayment months share the same year numbering
        year_label = (month - 1) // 12 + 1

        ws.cell(row=row, column=1, value=month).font = font(size=9, color="808080")
        ws.cell(row=row, column=2, value=year_label).font = font(size=9, color="808080")

        # Opening balance
        if month == 1:
            ws.cell(row=row, column=3, value=principal).font = FONT_FORMULA
        else:
            ws.cell(row=row, column=3,
                    value=f"=G{row - 1}").font = FONT_FORMULA

        # Principal repayment
        if month <= moratorium:
            ws.cell(row=row, column=4, value=0).font = FONT_FORMULA
        else:
            # EMI principal from Assumption — loan amount ÷ repayment months
            ws.cell(row=row, column=4,
                    value=f"={tl_amt_ref}/{tl_rep_ref}").font = FONT_FORMULA

        # Interest = opening × monthly rate
        ws.cell(row=row, column=5,
                value=f"=C{row}*{monthly_rate_formula}").font = FONT_FORMULA

        # Total EMI
        ws.cell(row=row, column=6, value=f"=D{row}+E{row}").font = FONT_FORMULA

        # Closing = opening - principal
        ws.cell(row=row, column=7, value=f"=C{row}-D{row}").font = FONT_FORMULA

        for col in [3, 4, 5, 6, 7]:
            ws.cell(row=row, column=col).number_format = FMT_LAKHS

    # ── Annual summary (right side of sheet) ───────────────────────────────────
    ANN_COL_START = 9   # column I
    ANN_ROW_LABEL = SCHED_START - 1
    ANN_ROW_YEAR  = SCHED_START
    ANN_PRINCIPAL = SCHED_START + 1
    ANN_CLOSING   = SCHED_START + 2
    ANN_INTEREST  = SCHED_START + 3

    for col, label in [(ANN_COL_START + i, f"Year {i+1}")
                       for i in range(layout.n_years)]:
        c = ws.cell(row=ANN_ROW_LABEL, column=col, value=label)
        c.font = font(bold=True, color="FFFFFF")
        c.fill = fill("1F3864")
        c.alignment = AL_CENTER

    ws.cell(row=ANN_ROW_YEAR,  column=ANN_COL_START - 1,
            value="Annual Principal Repaid").font = font(bold=True, size=9)
    ws.cell(row=ANN_CLOSING,   column=ANN_COL_START - 1,
            value="Year-End Outstanding").font = font(bold=True, size=9)
    ws.cell(row=ANN_INTEREST,  column=ANN_COL_START - 1,
            value="Annual Interest Paid").font = font(bold=True, size=9)

    sched_first_row = SCHED_START
    sched_last_row  = SCHED_START + tenor - 1

    for y in range(1, layout.n_years + 1):
        col   = ANN_COL_START + y - 1
        col_l = get_column_letter(col)

        # Annual principal = SUMIF on year column (B) = year label
        ws.cell(row=ANN_ROW_YEAR, column=col,
                value=(f"=SUMIF($B${sched_first_row}:$B${sched_last_row},"
                       f"{y},$D${sched_first_row}:$D${sched_last_row})")).font = FONT_FORMULA

        # Year-end outstanding — find last closing balance for this year
        ws.cell(row=ANN_CLOSING, column=col,
                value=(f"=IFERROR(INDEX($G${sched_first_row}:$G${sched_last_row},"
                       f"MATCH({y},$B${sched_first_row}:$B${sched_last_row},0)"
                       f"+COUNTIF($B${sched_first_row}:$B${sched_last_row},{y})-1),0)")).font = FONT_FORMULA

        # Annual interest
        ws.cell(row=ANN_INTEREST, column=col,
                value=(f"=SUMIF($B${sched_first_row}:$B${sched_last_row},"
                       f"{y},$E${sched_first_row}:$E${sched_last_row})")).font = FONT_FORMULA

        for row in [ANN_ROW_YEAR, ANN_CLOSING, ANN_INTEREST]:
            ws.cell(row=row, column=col).number_format = FMT_LAKHS

    return ws
