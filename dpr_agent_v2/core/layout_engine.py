"""
core/layout_engine.py
══════════════════════
Computes Excel cell addresses dynamically from the assumption registry.

KEY PRINCIPLE: No row number is hardcoded anywhere in any sheet writer.
Every address is computed here via:
    row = section_anchor + (item_index × rows_per_item) + field_offset

Usage:
    layout = LayoutEngine(n_products=3, n_materials=5, n_employees=5)

    layout.asmp_row("capacity.year1_util")          → 4
    layout.asmp_row("revenue.product.price", idx=0) → 15
    layout.asmp_row("revenue.product.price", idx=2) → 23
    layout.asmp_row("material.price", idx=0)        → 58
    layout.asmp_row("opex.power_pct_revenue")       → 123
    layout.asmp_addr("opex.rm_pct_fa")              → "E119"
"""

from typing import Optional
from openpyxl.utils import get_column_letter

from core.assumption_registry import (
    FieldDef, Section, FieldTier,
    SECTION_ANCHOR, FIELDS, get_field,
    ANCHOR_K,
    MAX_PRODUCTS, MAX_MATERIALS, MAX_EMPLOYEES,
    ROWS_PER_PRODUCT, ROWS_PER_MATERIAL, ROWS_PER_EMPLOYEE,
)


# Column assignments in the Assumption sheet
COL_ID     = 1   # A — field ID (e.g. A1, B3, C2)
COL_LABEL  = 2   # B — human label
COL_BLANK  = 3   # C — spacer
COL_UNIT   = 4   # D — unit
COL_VALUE  = 5   # E — the actual value (blue=T1, green=T2, grey=T3)
COL_AUX    = 6   # F — auxiliary value (split%, input_per_output, increment)


class LayoutEngine:
    """
    Computes row numbers and Excel addresses for every cell in the
    Assumption sheet, and the corresponding column start rows for
    all downstream sheets (Revenue, Expenses, PL, BS, etc.).

    Parameters
    ----------
    n_products   : actual number of products (1–MAX_PRODUCTS)
    n_materials  : actual number of raw materials (1–MAX_MATERIALS)
    n_employees  : actual number of employee categories (1–MAX_EMPLOYEES)
    n_years      : projection years (default 7)
    """

    def __init__(
        self,
        n_products:  int = 1,
        n_materials: int = 1,
        n_employees: int = 1,
        n_years:     int = 7,
    ):
        assert 1 <= n_products  <= MAX_PRODUCTS,  f"n_products {n_products} out of range"
        assert 1 <= n_materials <= MAX_MATERIALS, f"n_materials {n_materials} out of range"
        assert 1 <= n_employees <= MAX_EMPLOYEES, f"n_employees {n_employees} out of range"
        assert 1 <= n_years     <= 15,            f"n_years {n_years} out of range"

        self.n_products  = n_products
        self.n_materials = n_materials
        self.n_employees = n_employees
        self.n_years     = n_years

    # ── Core row computation ──────────────────────────────────────────────────

    def asmp_row(self, key: str, idx: int = 0) -> int:
        """
        Return the row number for a field in the Assumption sheet.

        Parameters
        ----------
        key : field key from assumption_registry (e.g. "material.price")
        idx : item index for repeating fields (product 0,1,2 / material 0,1 etc.)
        """
        f = get_field(key)
        anchor = SECTION_ANCHOR[f.section]

        if f.is_repeating:
            # Row = anchor + header_rows + (item_index × rows_per_item) + field_offset
            # anchor points to the section header row; data starts at anchor+1
            return anchor + 1 + (idx * f.rows_per_item) + f.item_field_offset
        else:
            # Fixed field — just anchor + offset
            return anchor + f.row_offset

    def asmp_addr(self, key: str, idx: int = 0, col: int = COL_VALUE) -> str:
        """
        Return Excel address string for a field e.g. "$E$47".
        Always returns an absolute reference ($ prefix).
        """
        row = self.asmp_row(key, idx)
        return f"$E${row}" if col == COL_VALUE else f"${get_column_letter(col)}${row}"

    def asmp_ref(self, key: str, idx: int = 0, col: int = COL_VALUE) -> str:
        """
        Return a cross-sheet formula reference e.g. "Assumption!$E$47".
        Use this in all sheet formula strings.
        """
        return f"Assumption!{self.asmp_addr(key, idx, col)}"

    # ── Revenue sheet layout ──────────────────────────────────────────────────
    # Revenue sheet: data starts at column D (Year 1), each year one column right
    # Row structure per product:
    #   base_row + 0 : section header (product name)
    #   base_row + 1 : total production (units)
    #   base_row + 2 : price per unit
    #   base_row + 3 : revenue (lakhs)

    REV_HEADER_ROW    = 4    # year headers
    REV_PARAMS_START  = 5    # rows 5–10: months, working days, capacity, utilisation
    REV_UTIL_ROW      = 10   # capacity utilisation row
    REV_DATA_START    = 12   # product blocks start here
    REV_ROWS_PER_PROD = 6    # header + production + liters + net_output + price + revenue
    REV_COL_YEAR1     = 4    # column D = Year 1

    def rev_product_start(self, prod_idx: int) -> int:
        """First row of a product block in the Revenue sheet."""
        return self.REV_DATA_START + (prod_idx * self.REV_ROWS_PER_PROD)

    def rev_production_row(self, prod_idx: int) -> int:
        return self.rev_product_start(prod_idx) + 1

    def rev_price_row(self, prod_idx: int) -> int:
        return self.rev_product_start(prod_idx) + 4

    def rev_revenue_row(self, prod_idx: int) -> int:
        return self.rev_product_start(prod_idx) + 5

    def rev_total_row(self) -> int:
        """TOTAL REVENUE row — after all product blocks."""
        return self.REV_DATA_START + (self.n_products * self.REV_ROWS_PER_PROD) + 2

    def year_col(self, year: int) -> int:
        """Column index for Year N (1-based). Year 1 = col D = 4."""
        return self.REV_COL_YEAR1 + (year - 1)

    def year_col_letter(self, year: int) -> str:
        return get_column_letter(self.year_col(year))

    # ── Expenses sheet layout ─────────────────────────────────────────────────
    # Each material occupies 3 rows: quantity, price, cost
    EXP_HEADER_ROW     = 1
    EXP_COGS_START     = 5    # raw material blocks start
    EXP_ROWS_PER_MAT   = 3   # quantity, price, cost
    EXP_COL_YEAR1      = 4   # column D

    def exp_material_start(self, mat_idx: int) -> int:
        return self.EXP_COGS_START + (mat_idx * self.EXP_ROWS_PER_MAT)

    def exp_qty_row(self, mat_idx: int) -> int:
        return self.exp_material_start(mat_idx)

    def exp_price_row(self, mat_idx: int) -> int:
        return self.exp_material_start(mat_idx) + 1

    def exp_cost_row(self, mat_idx: int) -> int:
        return self.exp_material_start(mat_idx) + 2

    def exp_total_cogs_row(self) -> int:
        """TOTAL COST OF SALES row."""
        return self.EXP_COGS_START + (self.n_materials * self.EXP_ROWS_PER_MAT) + 2

    def exp_overhead_start(self) -> int:
        """Where overhead expense rows begin."""
        return self.exp_total_cogs_row() + 3

    # Overhead row offsets from exp_overhead_start()
    EXP_OH_RM_RATE          = 0
    EXP_OH_RM_AMOUNT        = 1
    EXP_OH_INS_RATE         = 2
    EXP_OH_INS_AMOUNT       = 3
    EXP_OH_POWER_RATE       = 4
    EXP_OH_POWER_AMOUNT     = 5
    EXP_OH_MARKETING_RATE   = 6
    EXP_OH_MARKETING_AMOUNT = 7
    EXP_OH_SGA_BASE         = 8
    EXP_OH_SGA_AMOUNT       = 9
    EXP_OH_TRANSPORT_BASE   = 10
    EXP_OH_TRANSPORT_AMOUNT = 11
    EXP_OH_MISC_BASE        = 12
    EXP_OH_MISC_AMOUNT      = 13

    def exp_oh_row(self, offset: int) -> int:
        return self.exp_overhead_start() + offset

    def exp_total_oh_row(self) -> int:
        return self.exp_oh_row(14) + 1

    # ── ManPower sheet layout ─────────────────────────────────────────────────
    MP_DATA_START    = 7    # employee data rows start
    MP_ROWS_PER_EMP  = 1   # one row per employee category
    MP_COL_YEAR1     = 4   # column D

    def mp_employee_row(self, emp_idx: int) -> int:
        return self.MP_DATA_START + emp_idx

    def mp_base_total_row(self) -> int:
        return self.MP_DATA_START + self.n_employees + 1

    def mp_annual_row(self) -> int:
        """Annual salary projection row (escalated years)."""
        return self.mp_base_total_row() + 3

    # ── Depreciation sheet layout ─────────────────────────────────────────────
    DEP_ASSET_START  = 5
    DEP_COL_YEAR1    = 4

    # Asset type offsets (each asset = 3 rows: opening, depr, closing)
    DEP_ROWS_PER_ASSET = 3

    # Asset order: PM, Civil, Furniture, Vehicle, Electrical, Other
    ASSET_ORDER = ["plant_machinery", "civil_works", "furniture",
                   "vehicle", "electrical", "other"]

    def dep_asset_row(self, asset_type: str, row_type: str = "depr") -> int:
        """
        row_type: "opening" | "depr" | "closing"
        """
        idx = self.ASSET_ORDER.index(asset_type)
        base = self.DEP_ASSET_START + (idx * self.DEP_ROWS_PER_ASSET)
        offsets = {"opening": 0, "depr": 1, "closing": 2}
        return base + offsets[row_type]

    def dep_total_depr_row(self) -> int:
        """Total depreciation charge row."""
        return self.DEP_ASSET_START + (len(self.ASSET_ORDER) * self.DEP_ROWS_PER_ASSET) + 2

    def dep_net_block_row(self) -> int:
        """Net Fixed Assets (WDV) row — used by Expenses and BS."""
        return self.dep_total_depr_row() + 1

    # ── Term Loan sheet layout ────────────────────────────────────────────────
    TL_SCHEDULE_START = 5   # monthly schedule starts here
    TL_ANNUAL_START   = 5   # annual summary (right side of sheet)
    TL_ANNUAL_COL_Y1  = 9   # column I = Year 1 annual summary

    def tl_annual_col(self, year: int) -> int:
        return self.TL_ANNUAL_COL_Y1 + (year - 1)

    # Annual summary rows (relative to TL_ANNUAL_START)
    TL_ANNUAL_PRINCIPAL_ROW = 5   # SUMIF is at SCHED_START (row 5)
    TL_ANNUAL_CLOSING_ROW   = 7
    TL_ANNUAL_INTEREST_ROW  = 8

    # ── P&L sheet layout ──────────────────────────────────────────────────────
    # PL data columns must align with Revenue sheet: Year 1 = col D
    PL_COL_YEAR1 = 4   # column D — MUST match revenue sheet

    PL_REVENUE_ROW        = 9
    PL_TOTAL_REVENUE_ROW  = 10
    PL_COGS_ROW           = 13
    PL_GROSS_PROFIT_ROW   = 16
    PL_OPEX_START_ROW     = 19
    PL_RM_ROW             = 20
    PL_INS_ROW            = 21
    PL_MARKETING_ROW      = 22
    PL_POWER_ROW          = 23
    PL_MANPOWER_ROW       = 24
    PL_DEPR_ROW           = 25
    PL_SGA_ROW            = 26
    PL_TRANSPORT_ROW      = 27
    PL_MISC_ROW           = 28
    PL_TOTAL_OPEX_ROW     = 29
    PL_EBIT_ROW           = 31
    PL_TL_INTEREST_ROW    = 34
    PL_WC_INTEREST_ROW    = 35
    PL_TOTAL_FINANCE_ROW  = 36
    PL_PBT_ROW            = 38
    PL_TAX_ROW            = 40
    PL_PAT_ROW            = 42
    PL_DRAWINGS_ROW       = 43   # Less: Drawings / Dividends
    PL_RETAINED_ROW       = 44   # Net Retained (PAT - Drawings)
    PL_EBITDA_ROW         = 47
    PL_EBITDA_MARGIN_ROW  = 48   # EBITDA / Revenue
    PL_ICR_ROW            = 49   # Interest Coverage (EBIT / Finance Costs)

    def pl_col(self, year: int) -> int:
        return self.PL_COL_YEAR1 + (year - 1)

    def pl_col_letter(self, year: int) -> str:
        return get_column_letter(self.pl_col(year))

    # ── Balance Sheet layout ──────────────────────────────────────────────────
    BS_COL_YEAR1 = 3   # column C

    # ── LIABILITIES ───────────────────────────────────────────────────────────
    BS_TL_ROW               = 6    # Term Loan outstanding
    BS_VEHICLE_LOAN_ROW     = 7    # Vehicle Loan (default 0)
    BS_UNSECURED_LOANS_ROW  = 8    # Unsecured Loans (default 0)
    BS_OTHER_TERM_LIAB_ROW  = 9    # Other Term Liabilities (default 0)
    BS_TOTAL_TERM_LIAB_ROW  = 10   # Total Term Liabilities
    BS_TRADE_CRED_ROW       = 11   # Trade Creditors (WC)
    BS_OD_LOAN_ROW          = 12   # WC Loan / OD Utilised
    BS_TOTAL_CL_ROW         = 13   # Total Current Liabilities
    BS_TOTAL_OUTSIDE_LIAB   = 14   # Total Outside Liabilities (Term + Current)
    BS_EQUITY_ROW           = 16   # Share Capital
    BS_RESERVES_ROW         = 17   # Reserves & Surplus
    BS_TOTAL_EQUITY_ROW     = 18   # Shareholders' Fund
    BS_TOTAL_LIAB_ROW       = 19   # Total Liabilities (Outside + Equity)

    # ── ASSETS ────────────────────────────────────────────────────────────────
    BS_CASH_ROW             = 23   # Cash & Bank Balance
    BS_FD_ROW               = 24   # Fixed Deposits with Banks
    BS_DEBTORS_ROW          = 25   # Trade Receivables (Debtors)
    BS_CONSUMABLES_ROW      = 27   # Inventory: Consumables (RM Stock)
    BS_WIP_ROW              = 28   # WIP (default 0)
    BS_FG_ROW               = 29   # Finished Goods (default 0)
    BS_COLD_STORE_ROW       = 30   # Cold Store & Other (default 0)
    BS_TOTAL_CA_ROW         = 31   # Total Current Assets
    BS_GROSS_BLOCK_ROW      = 33   # Gross Block (original cost, constant)
    BS_CUM_DEPR_ROW         = 34   # Cumulative Depreciation charged
    BS_NET_BLOCK_ROW        = 35   # Net Block (WDV) = Gross - Cum Depr
    BS_INTANGIBLE_ROW       = 37
    BS_NON_CURR_INV_ROW     = 38
    BS_SECURITY_DEP_ROW     = 39
    BS_OTHER_NC_ROW         = 40   # Other Non-Current Assets
    BS_TOTAL_ASSETS_ROW     = 41   # Total Assets = CA + Net Block + NC
    BS_BALANCE_CHECK_ROW    = 43

    # Backward-compatible aliases
    BS_OTHER_CL_ROW         = 12   # alias for OD_LOAN_ROW
    BS_NFA_ROW              = 35   # alias for NET_BLOCK_ROW
    BS_WC_CREDITORS_ROW     = 11   # alias for TRADE_CRED_ROW
    BS_TOTAL_LTL_ROW        = 10   # alias for TOTAL_TERM_LIAB_ROW

    # ── BS BOTTOM RATIOS ──────────────────────────────────────────────────────
    BS_TNW_ROW              = 45   # Tangible Net Worth
    BS_NWC_ROW              = 46   # Net Working Capital
    BS_CURR_RATIO_ROW       = 47   # Current Ratio
    BS_CURR_RATIO_NO_OD_ROW = 48   # Current Ratio excl OD
    BS_TOL_TNW_ROW          = 49   # Total Outside Liabilities / TNW

    def bs_col(self, year: int) -> int:
        return self.BS_COL_YEAR1 + (year - 1)

    def bs_col_letter(self, year: int) -> str:
        return get_column_letter(self.bs_col(year))

    # ── CFS layout ────────────────────────────────────────────────────────────
    CFS_COL_YEAR1 = 3

    CFS_PBT_ROW          = 5
    CFS_DEPR_ROW         = 6
    CFS_INTEREST_ROW     = 7
    CFS_WC_CHANGES_ROW   = 9    # Debtors=9, Stock=10, Creditors=11
    CFS_TAX_ROW          = 12
    CFS_DRAWINGS_ROW     = 13   # Less: Drawings / Proprietor Withdrawals
    CFS_NET_OPERATING    = 15
    CFS_CAPEX_ROW        = 18
    CFS_FD_ROW           = 19   # Investment & Deposits (FD etc.)
    CFS_OTHER_NC_ROW     = 20   # Other Non-Current Assets
    CFS_NON_OP_ROW       = 21   # Non-Operating Income
    CFS_NET_INVESTING    = 23
    CFS_TL_ROW           = 26   # Promoter Equity Brought In
    CFS_OD_ROW           = 28   # TL drawdown at 27 (TL_ROW+1), OD at 28
    CFS_INTEREST_PAID    = 29
    CFS_NET_FINANCING    = 31
    CFS_NET_CHANGE       = 33
    CFS_OPENING_CASH     = 34
    CFS_CLOSING_CASH     = 35

    def cfs_col(self, year: int) -> int:
        return self.CFS_COL_YEAR1 + (year - 1)

    # ── Ratio sheet layout ────────────────────────────────────────────────────
    RATIO_COL_YEAR1 = 3

    RATIO_DSCR_NUM_ROW  = 8
    RATIO_DSCR_DEN_ROW  = 10
    RATIO_DSCR_ROW      = 12
    RATIO_AVG_DSCR_ROW  = 13
    RATIO_ROCE_ROW      = 18
    RATIO_BEP_FIXED_ROW = 25
    RATIO_BEP_CONTRIB   = 27
    RATIO_BEP_ROW       = 29
    RATIO_OP_MARGIN       = 33
    RATIO_NET_MARGIN      = 34
    RATIO_DEBT_EQUITY     = 35
    RATIO_ASSET_TURNOVER  = 36   # Revenue / Total Assets
    RATIO_IRR_CF_ROW      = 39   # Free cash flows row (Y0..Y7) for IRR calc
    RATIO_IRR_ROW         = 40   # Project IRR

    def ratio_col(self, year: int) -> int:
        return self.RATIO_COL_YEAR1 + (year - 1)

    # ── W Cap sheet layout ────────────────────────────────────────────────────
    WCAP_COL_YEAR1 = 3

    WCAP_CREDITORS_RM    = 5
    WCAP_CREDITORS_ADMIN = 6
    WCAP_TOTAL_CL        = 7
    WCAP_STOCK_RM        = 10   # Raw material stock
    WCAP_STOCK_FG        = 11   # Finished goods stock
    WCAP_STOCK_WIP       = 12   # Work-in-progress
    WCAP_STOCK_COLD      = 13   # Cold store / other
    WCAP_DEBTORS         = 14   # Trade receivables
    WCAP_TOTAL_CA        = 16   # Total current assets
    WCAP_WC_REQUIREMENT  = 18
    WCAP_WC_LOAN         = 19
    WCAP_WC_INTEREST     = 20

    def wcap_col(self, year: int) -> int:
        return self.WCAP_COL_YEAR1 + (year - 1)

    # ── Utility helpers ───────────────────────────────────────────────────────

    def asmp_formula(self, key: str, idx: int = 0) -> str:
        """
        Return a formula string that references the Assumption sheet.
        e.g. "=Assumption!$E$47"
        """
        return f"={self.asmp_ref(key, idx)}"

    def all_year_cols(self) -> list[int]:
        """Return column indices for all projection years."""
        return [self.year_col(y) for y in range(1, self.n_years + 1)]

    def all_year_letters(self) -> list[str]:
        """Return column letters for all projection years."""
        return [self.year_col_letter(y) for y in range(1, self.n_years + 1)]

    def debug_summary(self) -> str:
        """Print a summary of key row addresses for debugging."""
        lines = [
            f"LayoutEngine(products={self.n_products}, materials={self.n_materials},"
            f" employees={self.n_employees}, years={self.n_years})",
            "",
            "=== ASSUMPTION SHEET KEY ROWS ===",
            f"  capacity.year1_util      → row {self.asmp_row('capacity.year1_util')}",
            f"  revenue.product.price p0 → row {self.asmp_row('revenue.product.price', 0)}",
            f"  revenue.product.price p1 → row {self.asmp_row('revenue.product.price', 1)}",
            f"  material.price m0        → row {self.asmp_row('material.price', 0)}",
            f"  material.price m1        → row {self.asmp_row('material.price', 1)}",
            f"  opex.rm_pct_fa           → row {self.asmp_row('opex.rm_pct_fa')}",
            f"  opex.power_pct_revenue   → row {self.asmp_row('opex.power_pct_revenue')}",
            f"  opex.transport_base      → row {self.asmp_row('opex.transport_base')}",
            f"  employee.salary e0       → row {self.asmp_row('employee.salary', 0)}",
            f"  employee.salary e1       → row {self.asmp_row('employee.salary', 1)}",
            f"  finance.tl_amount        → row {self.asmp_row('finance.tl_amount')}",
            f"  wc.debtor_days           → row {self.asmp_row('wc.debtor_days')}",
            f"  depr.plant_machinery     → row {self.asmp_row('depr.plant_machinery')}",
            f"  impl.months              → row {self.asmp_row('impl.months')}",
            f"  tax.entity_type          → row {self.asmp_row('tax.entity_type')}",
            "",
            "=== REVENUE SHEET ===",
            f"  product 0 revenue row    → {self.rev_revenue_row(0)}",
            f"  product 1 revenue row    → {self.rev_revenue_row(1)}",
            f"  total revenue row        → {self.rev_total_row()}",
            "",
            "=== EXPENSES SHEET ===",
            f"  material 0 cost row      → {self.exp_cost_row(0)}",
            f"  material 1 cost row      → {self.exp_cost_row(1)}",
            f"  total COGS row           → {self.exp_total_cogs_row()}",
            f"  R&M amount row           → {self.exp_oh_row(self.EXP_OH_RM_AMOUNT)}",
            f"  power amount row         → {self.exp_oh_row(self.EXP_OH_POWER_AMOUNT)}",
        ]
        return "\n".join(lines)
